"""
apps/advanced_psychotest/views.py

Add-On: Advanced Psychometric Test Suite
Views untuk HR (manajemen sesi) dan kandidat (mengerjakan tes).
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count, Avg
from django.core.paginator import Paginator

from .models import AdvSoal, AdvSession, AdvAnswer, AdvResult, TEST_TYPE_CHOICES, TEST_DURATION
from apps.recruitment.models import Candidate
from apps.core.addon_decorators import addon_required

def _get_company(request):
    """Helper multi-tenant: ambil company aktif dari request."""
    company = getattr(request, 'company', None)
    if not company and getattr(request.user, 'is_superuser', False):
        from apps.core.models import Company
        company = Company.objects.first()
    return company



ADDON_KEY = 'advanced_psychotest'


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _compute_score(session):
    """Hitung skor dan buat / update AdvResult setelah kandidat submit."""
    answers = session.answers.select_related('soal').all()
    test_type = session.test_type

    result, _ = AdvResult.objects.get_or_create(
        session=session,
        defaults={
            'candidate': session.candidate,
            'employee':  session.employee,
            'test_type': test_type,
        }
    )

    if test_type in ('raven', 'cogspeed', 'cfit'):
        total_soal = answers.count()
        benar = sum(1 for a in answers if a.is_correct)
        skor = round(benar / total_soal * 100) if total_soal else 0
        result.skor_total = skor
        result.detail = {'benar': benar, 'total': total_soal}

    elif test_type == 'sjt':
        total_poin = 0
        max_poin = 0
        detail = {}
        for a in answers:
            skor_map = {
                'A': a.soal.sjt_skor_a, 'B': a.soal.sjt_skor_b,
                'C': a.soal.sjt_skor_c, 'D': a.soal.sjt_skor_d,
            }
            max_poin += max(a.soal.sjt_skor_a, a.soal.sjt_skor_b,
                            a.soal.sjt_skor_c, a.soal.sjt_skor_d)
            poin = skor_map.get(a.jawaban, 0)
            total_poin += poin
            detail[str(a.soal.nomor)] = {'jawaban': a.jawaban, 'poin': poin}
        skor = round(total_poin / max_poin * 100) if max_poin else 0
        result.skor_total = skor
        result.detail = detail

    elif test_type == 'bigfive':
        dim_scores = {'O': [], 'C': [], 'E': [], 'A': [], 'N': []}
        for a in answers:
            if not a.likert_val:
                continue
            dim = a.soal.bigfive_dimensi
            val = a.likert_val
            if a.soal.bigfive_reverse:
                val = 6 - val  # reverse scoring
            if dim in dim_scores:
                dim_scores[dim].append(val)

        def to100(vals):
            if not vals: return None
            return round((sum(vals) / len(vals) - 1) / 4 * 100)

        result.ocean_o = to100(dim_scores['O'])
        result.ocean_c = to100(dim_scores['C'])
        result.ocean_e = to100(dim_scores['E'])
        result.ocean_a = to100(dim_scores['A'])
        result.ocean_n = to100(dim_scores['N'])
        vals = [v for v in [result.ocean_o, result.ocean_c, result.ocean_e,
                             result.ocean_a, result.ocean_n] if v is not None]
        result.skor_total = round(sum(vals) / len(vals)) if vals else None
        result.detail = {
            'O': result.ocean_o, 'C': result.ocean_c,
            'E': result.ocean_e, 'A': result.ocean_a, 'N': result.ocean_n,
        }
        result.interpretasi = _bigfive_interpretasi(result)

    result.grade = result.compute_grade()

    # Percentile estimasi sederhana berdasarkan skor
    if result.skor_total is not None:
        s = result.skor_total
        if s >= 90: result.percentile = 95
        elif s >= 80: result.percentile = 80
        elif s >= 70: result.percentile = 65
        elif s >= 60: result.percentile = 50
        elif s >= 50: result.percentile = 35
        elif s >= 40: result.percentile = 20
        else: result.percentile = 10

    result.save()
    return result


def _bigfive_interpretasi(result):
    lines = []
    def level(v):
        if v is None: return "N/A"
        if v >= 70: return "Tinggi"
        if v >= 40: return "Sedang"
        return "Rendah"

    desc = {
        'O': {
            'Tinggi': "Kreatif, penasaran, terbuka terhadap pengalaman baru. Cocok untuk peran inovatif.",
            'Sedang': "Cukup adaptif antara rutinitas dan ide baru.",
            'Rendah': "Lebih suka stabilitas, rutinitas, dan pendekatan yang terbukti.",
        },
        'C': {
            'Tinggi': "Terorganisir, disiplin, dapat diandalkan. Sangat cocok untuk peran yang membutuhkan akurasi.",
            'Sedang': "Cukup terorganisir, bekerja baik dengan sedikit bimbingan.",
            'Rendah': "Fleksibel namun perlu pengawasan lebih dalam manajemen tugas.",
        },
        'E': {
            'Tinggi': "Energik, suka bersosialisasi, cocok untuk peran customer-facing atau kepemimpinan.",
            'Sedang': "Dapat bekerja baik dalam tim maupun mandiri.",
            'Rendah': "Introvert, lebih produktif bekerja mandiri dan analitis.",
        },
        'A': {
            'Tinggi': "Kooperatif, empatik, membangun hubungan kerja yang baik.",
            'Sedang': "Cukup kooperatif, mampu bernegosiasi.",
            'Rendah': "Kompetitif dan kritis, cocok untuk peran negosiasi atau audit.",
        },
        'N': {
            'Tinggi': "Rentan terhadap stres, perlu support system yang baik.",
            'Sedang': "Cukup stabil secara emosional.",
            'Rendah': "Sangat stabil secara emosional, tenang di bawah tekanan.",
        },
    }

    for dim, label in [('O','Openness'), ('C','Conscientiousness'),
                       ('E','Extraversion'), ('A','Agreeableness'), ('N','Neuroticism')]:
        val = getattr(result, f'ocean_{dim.lower()}')
        lvl = level(val)
        lines.append(f"• {label} ({lvl} — {val}%): {desc[dim][lvl]}")
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# HR VIEWS — MANAJEMEN SESI
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@addon_required(ADDON_KEY)
def session_list(request):
    """Daftar semua sesi advanced test — kandidat & karyawan."""
    test_filter   = request.GET.get('test_type', '')
    status_filter = request.GET.get('status', '')
    tujuan_filter = request.GET.get('tujuan', '')
    page          = request.GET.get('page', 1)

    qs = AdvSession.objects.select_related('candidate', 'employee').order_by('-created_at')
    if test_filter:
        from django.db.models import Q
        # paket adalah JSONField — di SQLite pakai icontains karena stored as string
        # di PostgreSQL paket__contains tetap bekerja
        qs = qs.filter(Q(test_type=test_filter) | Q(paket__icontains=test_filter))
    if status_filter:
        qs = qs.filter(status=status_filter)
    if tujuan_filter:
        qs = qs.filter(tujuan=tujuan_filter)

    # Auto-update status expired
    now = timezone.now()
    qs.filter(status__in=['pending', 'started'], expired_at__lt=now).update(status='expired')

    paginator = Paginator(qs, 25)
    sessions_page = paginator.get_page(page)

    from django.db.models import Q
    stats_qs = AdvSession.objects.values('test_type').annotate(
        total=Count('id'),
        selesai=Count('id', filter=Q(status='completed'))
    )
    stats_map = {s['test_type']: s for s in stats_qs}
    stats_cards = []
    for ttype, tlabel in TEST_TYPE_CHOICES:
        s = stats_map.get(ttype, {})
        stats_cards.append({'ttype': ttype, 'tlabel': tlabel,
                            'total': s.get('total', 0), 'selesai': s.get('selesai', 0)})

    return render(request, 'advanced_psychotest/session_list.html', {
        'sessions':          sessions_page,
        'test_type_choices': TEST_TYPE_CHOICES,
        'test_filter':       test_filter,
        'status_filter':     status_filter,
        'tujuan_filter':     tujuan_filter,
        'stats_cards':       stats_cards,
        'paginator':         paginator,
    })


@login_required
@addon_required('advanced_psychotest')
def session_create(request, candidate_pk):
    """HR buat sesi advanced test — bisa pilih satu atau lebih tipe tes."""
    candidate = get_object_or_404(Candidate, pk=candidate_pk)

    if request.method == 'POST':
        paket = request.POST.getlist('paket')   # list dari checkbox
        if not paket:
            messages.error(request, 'Pilih minimal satu tipe tes.')
            return redirect('adv_session_create', candidate_pk=candidate_pk)

        # Validasi semua tipe valid
        valid_types = dict(TEST_TYPE_CHOICES)
        paket = [p for p in paket if p in valid_types]
        if not paket:
            messages.error(request, 'Tipe tes tidak valid.')
            return redirect('adv_session_create', candidate_pk=candidate_pk)

        # Cek duplikat sesi aktif yang punya overlap paket
        existing = AdvSession.objects.filter(
            candidate=candidate,
            status__in=['pending', 'started']
        ).first()
        if existing:
            messages.warning(request,
                f'Kandidat sudah memiliki sesi aktif. Selesaikan atau hapus sesi tersebut dulu.')
            return redirect('candidate_detail', pk=candidate_pk)

        from datetime import timedelta
        expired_days = int(request.POST.get('expired_days', 7))

        # Durasi per tipe dari form
        durasi_per_tes = {}
        for tipe in paket:
            val = request.POST.get(f'durasi_{tipe}')
            durasi_per_tes[tipe] = int(val) if val else TEST_DURATION.get(tipe, 25)

        session = AdvSession.objects.create(
            candidate=candidate,
            paket=paket,
            test_type=paket[0],             # backward-compat
            durasi_per_tes=durasi_per_tes,
            durasi_menit=durasi_per_tes[paket[0]],
            expired_at=timezone.now() + timedelta(days=expired_days),
            created_by=request.user.get_full_name() or request.user.username,
        )

        link = request.build_absolute_uri(f'/advanced-test/tes/{session.token}/')
        paket_label = ', '.join(valid_types[p] for p in paket)

        # Kirim email notifikasi ke kandidat
        from .emails import kirim_email_sesi_kandidat
        email_terkirim = kirim_email_sesi_kandidat(session, request)
        email_info = f' Email notifikasi {"terkirim ke " + candidate.email if email_terkirim else "gagal dikirim (cek konfigurasi email)"}.'

        messages.success(request,
            f'Sesi [{paket_label}] berhasil dibuat. Link: {link}{email_info}')
        return redirect('candidate_detail', pk=candidate_pk)

    # Hitung jumlah soal per tipe untuk ditampilkan di form
    soal_counts = {
        ttype: AdvSoal.objects.filter(test_type=ttype, aktif=True).count()
        for ttype, _ in TEST_TYPE_CHOICES
    }

    return render(request, 'advanced_psychotest/session_create.html', {
        'candidate':         candidate,
        'test_type_choices': TEST_TYPE_CHOICES,
        'test_durations':    TEST_DURATION,
        'soal_counts':       soal_counts,
    })


@login_required
@addon_required('advanced_psychotest')
def session_detail(request, pk):
    """Detail sesi + hasil tes."""
    session = get_object_or_404(AdvSession, pk=pk)

    # Multi-tes: ambil semua result, tampilkan yg pertama sebagai default
    results = list(session.results.order_by('test_type'))
    result  = results[0] if results else None

    # Untuk backward-compat: jika sesi lama (single test_type), ambil result tipe itu
    if result is None and session.test_type:
        result = session.results.filter(test_type=session.test_type).first()

    answers   = session.answers.select_related('soal').order_by('soal__test_type', 'soal__nomor')
    link_full = request.build_absolute_uri(f'/advanced-test/tes/{session.token}/')

    paket     = session.get_paket()
    tipe_labels = dict(TEST_TYPE_CHOICES)
    result_map  = {r.test_type: r for r in results}
    paket_results = [
        {'tipe': t, 'label': tipe_labels.get(t, t), 'result': result_map.get(t)}
        for t in paket
    ]

    # Peserta bisa kandidat atau karyawan
    peserta      = session.get_peserta()
    peserta_url  = session.get_peserta_url()

    return render(request, 'advanced_psychotest/session_detail.html', {
        'session':       session,
        'result':        result,
        'results':       results,
        'paket_results': paket_results,
        'answers':       answers,
        'link_full':     link_full,
        'peserta':       peserta,
        'peserta_url':   peserta_url,
        'ocean_labels': {
            'O': 'Openness', 'C': 'Conscientiousness',
            'E': 'Extraversion', 'A': 'Agreeableness', 'N': 'Neuroticism',
        },
    })


@login_required
@addon_required('advanced_psychotest')
def soal_bank(request):
    """HR lihat dan kelola bank soal advanced."""
    test_filter = request.GET.get('test_type', '')
    qs = AdvSoal.objects.filter(aktif=True)
    if test_filter:
        qs = qs.filter(test_type=test_filter)

    count_cards = [
        {'ttype': t, 'tlabel': l,
         'jumlah': AdvSoal.objects.filter(test_type=t, aktif=True).count()}
        for t, l in TEST_TYPE_CHOICES
    ]

    return render(request, 'advanced_psychotest/soal_bank.html', {
        'soal_list':         qs,
        'test_type_choices': TEST_TYPE_CHOICES,
        'test_filter':       test_filter,
        'count_cards':       count_cards,
    })


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC VIEWS — KANDIDAT MENGERJAKAN TES
# ─────────────────────────────────────────────────────────────────────────────

def tes_intro(request, token):
    """Halaman intro — tampilkan checklist semua tes dalam paket."""
    session = get_object_or_404(AdvSession, token=token)

    if session.status == 'completed':
        return redirect('adv_tes_selesai', token=token)
    if session.is_expired:
        session.status = 'expired'
        session.save(update_fields=['status'])
        return render(request, 'advanced_psychotest/tes_expired.html', {'session': session})

    paket = session.get_paket()
    tes_selesai = session.get_tes_selesai()

    # Info per tes untuk ditampilkan di checklist
    INFO_TES = {
        'raven':    {'icon': '🔲', 'deskripsi': 'Penalaran visual & logika non-verbal'},
        'cogspeed': {'icon': '⚡', 'deskripsi': 'Kecepatan & ketepatan kognitif'},
        'bigfive':  {'icon': '⭐', 'deskripsi': 'Kepribadian OCEAN (tidak ada benar/salah)'},
        'sjt':      {'icon': '⚖️', 'deskripsi': 'Penilaian situasi kerja nyata'},
        'cfit':     {'icon': '🌐', 'deskripsi': 'Intelegensi minim bias budaya'},
    }

    paket_info = []
    tipe_labels = dict(TEST_TYPE_CHOICES)
    for tipe in paket:
        n_soal = AdvSoal.objects.filter(test_type=tipe, aktif=True).count()
        selesai = tipe in tes_selesai
        paket_info.append({
            'tipe':     tipe,
            'label':    tipe_labels.get(tipe, tipe),
            'durasi':   session.get_durasi(tipe),
            'n_soal':   n_soal,
            'selesai':  selesai,
            'icon':     INFO_TES.get(tipe, {}).get('icon', '📝'),
            'deskripsi':INFO_TES.get(tipe, {}).get('deskripsi', ''),
        })

    # Tes pertama yang belum selesai
    tes_aktif = next((t for t in paket if t not in tes_selesai), None)

    return render(request, 'advanced_psychotest/tes_intro.html', {
        'session':    session,
        'paket_info': paket_info,
        'tes_selesai':tes_selesai,
        'tes_aktif':  tes_aktif,
        'semua_selesai': session.semua_selesai(),
    })


def tes_mulai(request, token):
    """Tampilkan soal tes — satu tipe per halaman, navigasi antar tes via ?tipe=."""
    session = get_object_or_404(AdvSession, token=token)

    if session.status == 'completed':
        return redirect('adv_tes_selesai', token=token)
    if session.is_expired:
        return render(request, 'advanced_psychotest/tes_expired.html', {'session': session})

    if session.status == 'pending':
        session.status = 'started'
        session.started_at = timezone.now()
        session.save(update_fields=['status', 'started_at'])

    paket = session.get_paket()
    tes_selesai = session.get_tes_selesai()

    # Tipe tes aktif — dari query param, default ke tes pertama yang belum selesai
    tipe_aktif = request.GET.get('tipe', '')
    if not tipe_aktif or tipe_aktif not in paket:
        tipe_aktif = next((t for t in paket if t not in tes_selesai), paket[0] if paket else '')

    if not tipe_aktif:
        return redirect('adv_tes_selesai', token=token)

    soal_qs = AdvSoal.objects.filter(test_type=tipe_aktif, aktif=True).order_by('nomor')
    soal_list = list(soal_qs)
    total = len(soal_list)

    # Guard: soal belum di-seed
    if total == 0:
        tipe_labels = dict(TEST_TYPE_CHOICES)
        return render(request, 'advanced_psychotest/tes_soal.html', {
            'session': session, 'soal': None, 'soal_list': [], 'total': 0,
            'answered_ids': set(), 'answered_count': 0, 'sisa_detik': None,
            'next_idx': 0, 'is_last': True, 'error_kosong': True,
            'tipe_aktif': tipe_aktif,
            'tipe_label': tipe_labels.get(tipe_aktif, tipe_aktif),
        })

    answered_ids = set(
        session.answers.filter(soal__test_type=tipe_aktif).values_list('soal_id', flat=True)
    )

    soal_idx = int(request.GET.get('q', 0))
    if soal_idx >= total:
        soal_idx = total - 1

    if request.method == 'POST':
        soal_id  = int(request.POST.get('soal_id', 0))
        soal_obj = get_object_or_404(AdvSoal, pk=soal_id)
        jawaban  = request.POST.get('jawaban', '').strip().upper()
        likert_val = request.POST.get('likert_val')
        next_idx   = int(request.POST.get('next_idx', soal_idx + 1))

        answer, _ = AdvAnswer.objects.get_or_create(session=session, soal=soal_obj)
        answer.jawaban = jawaban
        if likert_val:
            answer.likert_val = int(likert_val)
        answer.save()

        # Cek apakah soal terakhir tipe ini atau diminta submit tipe ini
        answered_now = session.answers.filter(soal__test_type=tipe_aktif).count()
        tipe_selesai = answered_now >= total or request.POST.get('submit_tipe')

        if tipe_selesai:
            # Hitung skor untuk tipe ini
            _compute_score_for_tipe(session, tipe_aktif)

            # Cek apakah semua tipe dalam paket selesai
            if session.semua_selesai():
                session.status = 'completed'
                session.completed_at = timezone.now()
                session.save(update_fields=['status', 'completed_at'])
                return redirect('adv_tes_selesai', token=token)

            # Lanjut ke tipe berikutnya yang belum selesai
            tes_selesai_updated = session.get_tes_selesai()
            tes_berikutnya = next(
                (t for t in paket if t not in tes_selesai_updated), None
            )
            if tes_berikutnya:
                return redirect(f'/advanced-test/tes/{token}/mulai/?tipe={tes_berikutnya}')
            else:
                session.status = 'completed'
                session.completed_at = timezone.now()
                session.save(update_fields=['status', 'completed_at'])
                return redirect('adv_tes_selesai', token=token)

        return redirect(f'/advanced-test/tes/{token}/mulai/?tipe={tipe_aktif}&q={next_idx}')

    soal = soal_list[soal_idx] if soal_list else None
    answered_count = len(answered_ids)

    # Catat waktu mulai tipe ini jika belum tercatat (poin 14: timer akurat)
    session.catat_tipe_started(tipe_aktif)
    sisa_detik = session.get_sisa_detik(tipe_aktif)

    # Info navigasi paket
    tipe_labels = dict(TEST_TYPE_CHOICES)
    paket_nav = []
    for t in paket:
        paket_nav.append({
            'tipe':    t,
            'label':   tipe_labels.get(t, t),
            'selesai': t in tes_selesai,
            'aktif':   t == tipe_aktif,
        })

    return render(request, 'advanced_psychotest/tes_soal.html', {
        'session':        session,
        'soal':           soal,
        'soal_idx':       soal_idx,
        'soal_list':      soal_list,
        'total':          total,
        'answered_ids':   answered_ids,
        'answered_count': answered_count,
        'sisa_detik':     sisa_detik,
        'next_idx':       soal_idx + 1,
        'is_last':        soal_idx >= total - 1,
        'tipe_aktif':     tipe_aktif,
        'tipe_label':     tipe_labels.get(tipe_aktif, tipe_aktif),
        'paket_nav':      paket_nav,
        'token':          token,
        'error_kosong':   False,
    })


def _compute_score_for_tipe(session, tipe):
    """Hitung & simpan AdvResult untuk satu tipe tes dalam sesi."""
    answers = list(session.answers.select_related('soal').filter(soal__test_type=tipe))

    defaults = {
        'candidate': session.candidate,
        'employee':  session.employee,
    }
    result, _ = AdvResult.objects.update_or_create(
        session=session,
        test_type=tipe,
        defaults=defaults,
    )

    if tipe in ('raven', 'cogspeed', 'cfit'):
        total_soal = len(answers)
        benar = sum(1 for a in answers if a.is_correct)
        skor = round(benar / total_soal * 100) if total_soal else 0
        result.skor_total = skor
        result.detail = {'benar': benar, 'total': total_soal}

    elif tipe == 'sjt':
        total_poin = 0
        max_poin   = 0
        detail     = {}
        for a in answers:
            skor_map = {
                'A': a.soal.sjt_skor_a, 'B': a.soal.sjt_skor_b,
                'C': a.soal.sjt_skor_c, 'D': a.soal.sjt_skor_d,
            }
            mp = max(a.soal.sjt_skor_a, a.soal.sjt_skor_b,
                     a.soal.sjt_skor_c, a.soal.sjt_skor_d)
            max_poin   += mp
            poin        = skor_map.get(a.jawaban, 0)
            total_poin += poin
            detail[str(a.soal.nomor)] = {'jawaban': a.jawaban, 'poin': poin}
        skor = round(total_poin / max_poin * 100) if max_poin else 0
        result.skor_total = skor
        result.detail = detail

    elif tipe == 'bigfive':
        dim_scores = {'O': [], 'C': [], 'E': [], 'A': [], 'N': []}
        for a in answers:
            if not a.likert_val:
                continue
            dim = a.soal.bigfive_dimensi
            val = a.likert_val
            if a.soal.bigfive_reverse:
                val = 6 - val
            if dim in dim_scores:
                dim_scores[dim].append(val)

        def to100(vals):
            return round((sum(vals) / len(vals) - 1) / 4 * 100) if vals else None

        result.ocean_o = to100(dim_scores['O'])
        result.ocean_c = to100(dim_scores['C'])
        result.ocean_e = to100(dim_scores['E'])
        result.ocean_a = to100(dim_scores['A'])
        result.ocean_n = to100(dim_scores['N'])
        vals = [v for v in [result.ocean_o, result.ocean_c, result.ocean_e,
                             result.ocean_a, result.ocean_n] if v is not None]
        result.skor_total = round(sum(vals) / len(vals)) if vals else None
        result.detail = {'O': result.ocean_o, 'C': result.ocean_c,
                         'E': result.ocean_e, 'A': result.ocean_a, 'N': result.ocean_n}
        result.interpretasi = _bigfive_interpretasi(result)

    result.grade = result.compute_grade()
    if result.skor_total is not None:
        s = result.skor_total
        result.percentile = (95 if s >= 90 else 80 if s >= 80 else 65 if s >= 70
                             else 50 if s >= 60 else 35 if s >= 50 else 20 if s >= 40 else 10)
    result.save()
    return result


def tes_selesai(request, token):
    session = get_object_or_404(AdvSession, token=token)
    results = list(AdvResult.objects.filter(session=session).order_by('test_type'))
    paket   = session.get_paket()
    tipe_labels = dict(TEST_TYPE_CHOICES)

    # Map result per tipe
    result_map = {r.test_type: r for r in results}
    paket_results = [
        {'tipe': t, 'label': tipe_labels.get(t, t), 'result': result_map.get(t)}
        for t in paket
    ]

    return render(request, 'advanced_psychotest/tes_selesai.html', {
        'session':       session,
        'paket_results': paket_results,
        'results':       results,
    })

# ─────────────────────────────────────────────────────────────────────────────
# PSIKOTES BERKALA KARYAWAN
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@addon_required('advanced_psychotest')
def employee_session_create(request, employee_pk):
    """HR buat sesi advanced test berkala untuk karyawan."""
    from apps.employees.models import Employee
    employee = get_object_or_404(Employee, pk=employee_pk)

    TUJUAN_CHOICES = [
        ('berkala',  'Evaluasi Berkala'),
        ('promosi',  'Pertimbangan Promosi'),
        ('evaluasi', 'Evaluasi Kinerja'),
        ('lainnya',  'Lainnya'),
    ]

    if request.method == 'POST':
        paket = request.POST.getlist('paket')
        if not paket:
            messages.error(request, 'Pilih minimal satu tipe tes.')
            return redirect('adv_employee_session_create', employee_pk=employee_pk)

        valid_types = dict(TEST_TYPE_CHOICES)
        paket = [p for p in paket if p in valid_types]

        from datetime import timedelta
        expired_days  = int(request.POST.get('expired_days', 7))
        tujuan        = request.POST.get('tujuan', 'berkala')
        durasi_per_tes = {}
        for tipe in paket:
            val = request.POST.get(f'durasi_{tipe}')
            durasi_per_tes[tipe] = int(val) if val else TEST_DURATION.get(tipe, 25)

        session = AdvSession.objects.create(
            employee       = employee,
            candidate      = None,
            paket          = paket,
            test_type      = paket[0],
            tujuan         = tujuan,
            durasi_per_tes = durasi_per_tes,
            durasi_menit   = durasi_per_tes[paket[0]],
            expired_at     = timezone.now() + timedelta(days=expired_days),
            created_by     = request.user.get_full_name() or request.user.username,
        )
        link = request.build_absolute_uri(f'/advanced-test/tes/{session.token}/')
        paket_label = ', '.join(valid_types[p] for p in paket)

        from .emails import kirim_email_sesi_karyawan
        email_terkirim = kirim_email_sesi_karyawan(session, request)
        email_info = f' Email {"terkirim" if email_terkirim else "gagal dikirim"}.'

        messages.success(request,
            f'Sesi [{paket_label}] untuk {employee.nama} berhasil dibuat. Link: {link}{email_info}')
        return redirect('employee_detail', pk=employee_pk)

    soal_counts = {t: AdvSoal.objects.filter(test_type=t, aktif=True).count()
                   for t, _ in TEST_TYPE_CHOICES}

    return render(request, 'advanced_psychotest/employee_session_create.html', {
        'employee':          employee,
        'test_type_choices': TEST_TYPE_CHOICES,
        'test_durations':    TEST_DURATION,
        'soal_counts':       soal_counts,
        'tujuan_choices':    TUJUAN_CHOICES,
    })


@login_required
@addon_required('advanced_psychotest')
def employee_psychotest_report(request, employee_pk):
    """Laporan lengkap hasil psikotes berkala satu karyawan."""
    from apps.employees.models import Employee
    employee = get_object_or_404(Employee, pk=employee_pk)
    sessions = AdvSession.objects.filter(
        employee=employee, status='completed'
    ).prefetch_related('results').order_by('created_at')

    tipe_labels = dict(TEST_TYPE_CHOICES)

    # Susun timeline: per sesi, per tipe tes
    timeline = []
    for sess in sessions:
        result_map = {r.test_type: r for r in sess.results.all()}
        timeline.append({
            'session':    sess,
            'tujuan':     sess.tujuan,
            'tanggal':    sess.completed_at,
            'results':    result_map,
            'paket':      sess.get_paket(),
        })

    # Tren per tipe tes — untuk grafik garis
    tren = {}
    for tipe, _ in TEST_TYPE_CHOICES:
        poin = []
        for t in timeline:
            r = t['results'].get(tipe)
            if r and r.skor_total is not None:
                poin.append({
                    'tanggal': t['tanggal'].strftime('%d/%m/%y') if t['tanggal'] else '',
                    'skor':    r.skor_total,
                    'grade':   r.grade,
                    'tujuan':  t['tujuan'],
                })
        if poin:
            tren[tipe] = poin

    # Skor terakhir per tipe
    skor_terakhir = {}
    for tipe in tren:
        skor_terakhir[tipe] = tren[tipe][-1]

    return render(request, 'advanced_psychotest/employee_report.html', {
        'employee':      employee,
        'timeline':      timeline,
        'tren':          tren,
        'tren_json':     __import__('json').dumps(tren),
        'skor_terakhir': skor_terakhir,
        'tipe_labels':   tipe_labels,
        'ocean_labels':  {'O':'Openness','C':'Conscientiousness','E':'Extraversion',
                          'A':'Agreeableness','N':'Neuroticism'},
    })


@login_required
@addon_required('advanced_psychotest')
def psychotest_report_all(request):
    """Dashboard laporan psikotes berkala seluruh karyawan."""
    from apps.employees.models import Employee
    from django.db.models import Avg, Count, Q

    dept_filter = request.GET.get('dept', '')
    tipe_filter = request.GET.get('tipe', 'bigfive')

    company = _get_company(request)
    employees = Employee.objects.filter(company=company, status='Aktif').select_related('department', 'jabatan') if company else Employee.objects.filter(status='Aktif').select_related('department', 'jabatan')
    if dept_filter:
        employees = employees.filter(department_id=dept_filter)

    # Ambil skor terakhir per karyawan per tipe tes
    profil_list = []
    for emp in employees:
        last_results = {}
        for tipe, _ in TEST_TYPE_CHOICES:
            r = AdvResult.objects.filter(
                session__employee=emp,
                test_type=tipe,
                session__status='completed',
            ).order_by('-created_at').first()
            if r:
                last_results[tipe] = r

        if last_results:
            profil_list.append({'employee': emp, 'results': last_results})

    # Rata-rata per departemen untuk tipe yang dipilih
    dept_avg = {}
    dept_qs = AdvResult.objects.filter(
        session__employee__isnull=False,
        test_type=tipe_filter,
        session__status='completed',
        skor_total__isnull=False,
    ).values(
        'session__employee__department__nama'
    ).annotate(avg_skor=Avg('skor_total'), total=Count('id'))

    for d in dept_qs:
        nama = d['session__employee__department__nama'] or 'Tanpa Dept'
        dept_avg[nama] = {'avg': round(d['avg_skor']), 'total': d['total']}

    from apps.core.models import Department
    company = _get_company(request)
    departments = Department.objects.filter(company=company, aktif=True) if company else Department.objects.filter(aktif=True)

    return render(request, 'advanced_psychotest/report_all.html', {
        'profil_list':       profil_list,
        'dept_avg':          dept_avg,
        'dept_avg_json':     __import__('json').dumps(dept_avg),
        'departments':       departments,
        'dept_filter':       dept_filter,
        'tipe_filter':       tipe_filter,
        'test_type_choices': TEST_TYPE_CHOICES,
        'tipe_labels':       dict(TEST_TYPE_CHOICES),
    })

# ─────────────────────────────────────────────────────────────────────────────
# EXPORT — Excel & PDF
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@addon_required('advanced_psychotest')
def export_excel(request):
    """
    Export semua hasil Advanced Test ke file Excel (.xlsx).
    Query param: ?tipe=bigfive&dept=1  (opsional, sama dengan report_all)
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Modul openpyxl belum terinstall. Jalankan: pip install openpyxl')
        return redirect('adv_report_all')

    from django.http import HttpResponse
    from apps.employees.models import Employee

    tipe_filter = request.GET.get('tipe', '')
    dept_filter = request.GET.get('dept', '')

    # Ambil semua hasil
    results_qs = AdvResult.objects.select_related(
        'session', 'session__employee', 'session__candidate',
        'session__employee__department', 'session__employee__jabatan',
    ).filter(session__status='completed')

    if tipe_filter:
        results_qs = results_qs.filter(test_type=tipe_filter)
    if dept_filter:
        results_qs = results_qs.filter(session__employee__department_id=dept_filter)

    results_qs = results_qs.order_by('test_type', '-created_at')

    # Buat workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hasil Advanced Test'

    # Style
    PURPLE = 'FF7B1FA2'
    LIGHT  = 'FFF3E5F5'
    HEADER_FONT  = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    HEADER_FILL  = PatternFill('solid', fgColor=PURPLE)
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    DATA_FONT    = Font(name='Calibri', size=10)
    ALT_FILL     = PatternFill('solid', fgColor='FFF5F5F5')
    BORDER_SIDE  = Side(style='thin', color='FFCCCCCC')
    CELL_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                          top=BORDER_SIDE, bottom=BORDER_SIDE)

    tipe_labels = dict(TEST_TYPE_CHOICES)

    # Header
    headers = [
        'No', 'Nama Peserta', 'Tipe', 'Jabatan/Posisi', 'Departemen',
        'Tujuan', 'Skor', 'Grade', 'Percentile',
        'O (Openness)', 'C (Conscient.)', 'E (Extraversion)',
        'A (Agreeableness)', 'N (Neuroticism)',
        'Tanggal Selesai',
    ]
    ws.row_dimensions[1].height = 36
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER

    # Data
    for i, r in enumerate(results_qs, 1):
        row_n = i + 1
        sess = r.session
        peserta = sess.get_peserta_nama()
        jabatan = ''
        dept    = ''
        if sess.employee:
            jabatan = str(sess.employee.jabatan) if sess.employee.jabatan else ''
            dept    = str(sess.employee.department) if sess.employee.department else ''
        elif sess.candidate:
            jabatan = getattr(sess.candidate, 'jabatan_dilamar', '') or ''

        fill = ALT_FILL if i % 2 == 0 else None
        row_data = [
            i, peserta,
            tipe_labels.get(r.test_type, r.test_type),
            jabatan, dept,
            sess.tujuan or 'recruitment',
            r.skor_total, r.grade, r.percentile,
            r.ocean_o, r.ocean_c, r.ocean_e, r.ocean_a, r.ocean_n,
            sess.completed_at.strftime('%d/%m/%Y %H:%M') if sess.completed_at else '',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_n, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill

    # Lebar kolom
    col_widths = [5, 24, 18, 20, 18, 12, 8, 8, 10, 10, 10, 10, 10, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    # Response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from django.utils import timezone as tz
    filename = f'adv_psychotest_{tz.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@addon_required('advanced_psychotest')
def export_session_detail_pdf(request, pk):
    """
    Export detail sesi satu kandidat/karyawan ke HTML yang bisa di-print/save as PDF.
    Menggunakan browser print dialog (tidak perlu WeasyPrint).
    """
    session = get_object_or_404(AdvSession, pk=pk)
    results = list(session.results.order_by('test_type'))
    paket   = session.get_paket()
    tipe_labels = dict(TEST_TYPE_CHOICES)
    result_map  = {r.test_type: r for r in results}
    paket_results = [
        {'tipe': t, 'label': tipe_labels.get(t, t), 'result': result_map.get(t)}
        for t in paket
    ]
    peserta     = session.get_peserta()
    peserta_url = session.get_peserta_url()

    return render(request, 'advanced_psychotest/session_print.html', {
        'session':       session,
        'results':       results,
        'paket_results': paket_results,
        'peserta':       peserta,
        'peserta_url':   peserta_url,
        'ocean_labels': {
            'O': 'Openness', 'C': 'Conscientiousness',
            'E': 'Extraversion', 'A': 'Agreeableness', 'N': 'Neuroticism',
        },
        'print_mode': True,
    })
