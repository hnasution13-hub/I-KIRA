# ==================================================
# FILE: apps/reports/export.py
# DESKRIPSI: Fungsi ekspor data ke Excel dan PDF
# VERSION: 1.0.0
# ==================================================

import io
from django.http import HttpResponse
from django.utils import timezone


def export_assets_excel(queryset):
    """Export daftar aset ke file Excel (.xlsx)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daftar Aset"

    # Style header
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Judul
    ws.merge_cells('A1:N1')
    ws['A1'] = f'DAFTAR ASET - {timezone.now().strftime("%d %B %Y")}'
    ws['A1'].font = Font(bold=True, size=13, color="1E3A5F")
    ws['A1'].alignment = center

    # Header kolom
    headers = [
        'No', 'Kode Aset', 'Nama Aset', 'Kategori', 'Merk/Brand',
        'No. Seri', 'Tgl Perolehan', 'Harga Perolehan', 'Masa Manfaat (Thn)',
        'Nilai Sisa', 'Lokasi', 'PIC', 'Status', 'Kondisi'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # Lebar kolom
    col_widths = [5, 15, 35, 20, 15, 18, 14, 20, 15, 15, 20, 20, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Data
    row_fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for idx, asset in enumerate(queryset, 1):
        row = idx + 3
        fill = row_fill_even if idx % 2 == 0 else None
        data = [
            idx,
            asset.asset_code,
            asset.asset_name,
            asset.category.name if asset.category else '-',
            asset.brand or '-',
            asset.serial_number or '-',
            asset.purchase_date.strftime('%d/%m/%Y') if asset.purchase_date else '-',
            float(asset.purchase_price),
            asset.useful_life,
            float(asset.residual_value),
            asset.location.name if asset.location else '-',
            asset.responsible.name if asset.responsible else '-',
            asset.status,
            asset.condition,
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if fill:
                cell.fill = fill
            if col in [1]:
                cell.alignment = center
            if col == 8:  # Harga
                cell.number_format = '#,##0'

    # Freeze panes
    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_assets_pdf(queryset):
    """Export daftar aset ke PDF menggunakan reportlab"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        return None

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                  fontSize=14, alignment=TA_CENTER, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'],
                                     fontSize=9, alignment=TA_CENTER, spaceAfter=20,
                                     textColor=colors.grey)

    content = []
    content.append(Paragraph('LAPORAN DAFTAR ASET', title_style))
    content.append(Paragraph(f'Tanggal: {timezone.now().strftime("%d %B %Y %H:%M")} WIB', subtitle_style))
    content.append(Spacer(1, 0.3*cm))

    # Header tabel
    table_data = [['No', 'Kode Aset', 'Nama Aset', 'Kategori', 'Tgl Perolehan',
                   'Harga (Rp)', 'Lokasi', 'PIC', 'Status', 'Kondisi']]

    for idx, asset in enumerate(queryset, 1):
        table_data.append([
            str(idx),
            asset.asset_code,
            asset.asset_name[:35],
            asset.category.name[:15] if asset.category else '-',
            asset.purchase_date.strftime('%d/%m/%Y') if asset.purchase_date else '-',
            f"{float(asset.purchase_price):,.0f}",
            asset.location.name[:15] if asset.location else '-',
            (asset.responsible.name[:15] if asset.responsible else '-'),
            asset.status,
            asset.condition,
        ])

    col_widths = [0.8*cm, 3*cm, 7*cm, 3.5*cm, 2.5*cm, 3.5*cm, 3*cm, 3*cm, 2*cm, 2.5*cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Data rows
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))

    content.append(table)
    content.append(Spacer(1, 0.5*cm))

    # Footer
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'],
                                   fontSize=8, textColor=colors.grey)
    content.append(Paragraph(f'Total: {queryset.count()} aset | Diekspor pada {timezone.now().strftime("%d/%m/%Y %H:%M")}', footer_style))

    doc.build(content)
    output.seek(0)
    return output


def export_depreciation_excel(queryset):
    """Export laporan depresiasi ke Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Depresiasi"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A1:H1')
    ws['A1'] = f'LAPORAN DEPRESIASI ASET - {timezone.now().strftime("%d %B %Y")}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    headers = ['Kode Aset', 'Nama Aset', 'Tgl Perolehan', 'Harga Perolehan',
               'Masa Manfaat', 'Nilai Sisa', 'Depresiasi/Bulan', 'Nilai Buku Terakhir']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    for i in range(4, 9):
        ws.column_dimensions[get_column_letter(i)].width = 20

    for idx, asset in enumerate(queryset, 1):
        row = idx + 3
        ws.cell(row=row, column=1, value=asset.asset_code).border = thin_border
        ws.cell(row=row, column=2, value=asset.asset_name).border = thin_border
        ws.cell(row=row, column=3, value=asset.purchase_date.strftime('%d/%m/%Y')).border = thin_border
        c4 = ws.cell(row=row, column=4, value=float(asset.purchase_price))
        c4.number_format = '#,##0'
        c4.border = thin_border
        ws.cell(row=row, column=5, value=f'{asset.useful_life} tahun').border = thin_border
        c6 = ws.cell(row=row, column=6, value=float(asset.residual_value))
        c6.number_format = '#,##0'
        c6.border = thin_border
        c7 = ws.cell(row=row, column=7, value=float(asset.monthly_depreciation()))
        c7.number_format = '#,##0'
        c7.border = thin_border
        c8 = ws.cell(row=row, column=8, value=float(asset.book_value()))
        c8.number_format = '#,##0'
        c8.border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
