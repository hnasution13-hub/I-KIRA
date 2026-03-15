"""
apps/assets/export_import.py
=============================
Utilitas download template + import data aset dari Excel.
Konsisten dengan pola import karyawan & payroll.
"""
import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

C_NAVY   = '1E3A5F'
C_RED    = 'C0392B'
C_GREEN  = '1A7A4A'
C_ORANGE = 'E67E22'
C_YELLOW = 'FFF3CD'
C_WHITE  = 'FFFFFF'
C_GRAY   = 'F5F5F5'
C_LGRAY  = 'FAFAFA'


def _border(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Definisi kolom import ─────────────────────────────────────────────────────
# (label, keterangan, wajib, lebar, contoh)
IMPORT_COLUMNS = [
    ('Kode Aset',        'Kode unik aset, contoh: KND-R4-0001. Kosongkan = auto-generate.',  False, 16, 'KND-R4-0001'),
    ('Nama Aset',        'Nama lengkap aset',                                                  True,  30, 'Toyota Avanza 2023'),
    ('Kategori Induk',   'Nama kategori level 1. Akan dibuat otomatis jika belum ada.',        False, 20, 'Kendaraan'),
    ('Kategori',         'Nama kategori level 2 (sub-kategori). Akan dibuat otomatis.',        False, 20, 'Kendaraan Roda 4'),
    ('Tanggal Beli',     'Format: DD/MM/YYYY',                                                 True,  16, '15/01/2023'),
    ('Harga Beli',       'Angka saja tanpa titik/koma. Contoh: 150000000',                    True,  18, '150000000'),
    ('Masa Manfaat',     'Dalam tahun. Isi 0 untuk tanah (tidak disusutkan).',                False, 14, '8'),
    ('Nilai Residu',     'Angka saja. Contoh: 15000000. Default 0 jika kosong.',              False, 16, '15000000'),
    ('Merk / Brand',     'Merk atau brand aset',                                               False, 16, 'Toyota'),
    ('Serial Number',    'Nomor seri / nomor rangka / plat nomor',                             False, 18, 'B 1234 XYZ'),
    ('Kondisi',          'Baik / Rusak Ringan / Rusak Berat / Dalam Perbaikan',               False, 18, 'Baik'),
    ('Status',           'ACTIVE / MAINTENANCE / RETIRED / BROKEN / RESERVED',                False, 14, 'ACTIVE'),
    ('Lokasi',           'Nama lokasi / ruangan (harus sudah ada, atau akan dilewati).',       False, 22, 'Ruang HRD'),
    ('Vendor',           'Nama vendor / supplier (harus sudah ada, atau akan dilewati).',      False, 22, 'PT Sumber Teknologi'),
    ('Penanggung Jawab', 'NIK karyawan penanggung jawab (harus sudah ada di sistem).',         False, 20, 'EMP001'),
    ('Garansi s.d.',     'Tanggal akhir garansi. Format: DD/MM/YYYY. Kosongkan jika tidak ada.', False, 16, '15/01/2026'),
    ('No. Nota/Invoice', 'Nomor nota atau invoice pembelian',                                  False, 20, 'INV/2023/01/001'),
    ('Catatan',          'Catatan tambahan tentang aset',                                      False, 30, 'Kendaraan operasional direksi'),
]

REFERENSI = [
    ('Kondisi', ['Baik', 'Rusak Ringan', 'Rusak Berat', 'Dalam Perbaikan']),
    ('Status',  ['ACTIVE', 'MAINTENANCE', 'RETIRED', 'BROKEN', 'RESERVED']),
    ('Kategori Induk (Contoh)', ['Kendaraan', 'Properti', 'Peralatan Kantor', 'Infrastruktur IT', 'Mesin', 'AC & Elektronik']),
    ('Kategori (Contoh)',       ['Kendaraan Roda 4', 'Kendaraan Roda 2', 'Gedung & Bangunan', 'Tanah',
                                  'Komputer & Laptop', 'Printer & Scanner', 'Furniture',
                                  'Server & Storage', 'Jaringan', 'AC & Pendingin', 'Elektronik', 'Mesin & Peralatan']),
]


def download_template_import_asset(company=None):
    """Generate file Excel template import aset."""
    wb   = openpyxl.Workbook()
    ncol = len(IMPORT_COLUMNS)

    # ── Sheet 1: Template ─────────────────────────────────────────────────────
    ws       = wb.active
    ws.title = 'Template Import Aset'

    # Baris 1: Judul
    ws.merge_cells(f'A1:{get_column_letter(ncol)}1')
    c           = ws['A1']
    c.value     = 'TEMPLATE IMPORT DATA ASET — HRIS SmartDesk'
    c.font      = Font(bold=True, size=14, color=C_WHITE)
    c.fill      = PatternFill('solid', fgColor=C_NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # Baris 2: Info tenant
    ws.merge_cells(f'A2:{get_column_letter(ncol)}2')
    c           = ws['A2']
    tenant_nama = company.nama if company else '— (tenant dipilih saat upload) —'
    c.value     = f'🏢  Tenant: {tenant_nama}   |   Kategori & Lokasi akan dibuat otomatis jika belum ada'
    c.font      = Font(bold=True, size=10, color='1A5276')
    c.fill      = PatternFill('solid', fgColor='D6EAF8')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # Baris 3: Peringatan
    ws.merge_cells(f'A3:{get_column_letter(ncol)}3')
    c           = ws['A3']
    c.value     = '⚠  Kolom merah * = WAJIB  |  Format tanggal: DD/MM/YYYY  |  Harga: angka saja tanpa titik/koma  |  Hapus baris contoh (baris 6) sebelum import'
    c.font      = Font(bold=True, size=9, color=C_RED)
    c.fill      = PatternFill('solid', fgColor=C_YELLOW)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 20

    # Baris 4: Header kolom
    for col, (label, _, wajib, width, _) in enumerate(IMPORT_COLUMNS, 1):
        display     = f'{label} *' if wajib else label
        c           = ws.cell(row=4, column=col, value=display)
        c.font      = Font(bold=True, color=C_WHITE, size=10)
        c.fill      = PatternFill('solid', fgColor=C_RED if wajib else C_NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[4].height = 30

    # Baris 5: Keterangan
    for col, (_, ket, _, _, _) in enumerate(IMPORT_COLUMNS, 1):
        c           = ws.cell(row=5, column=col, value=ket)
        c.font      = Font(italic=True, size=8, color='666666')
        c.fill      = PatternFill('solid', fgColor=C_GRAY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[5].height = 40

    # Baris 6: Contoh (abu-abu)
    for col, (_, _, _, _, contoh) in enumerate(IMPORT_COLUMNS, 1):
        c           = ws.cell(row=6, column=col, value=contoh)
        c.font      = Font(italic=True, size=9, color='999999')
        c.fill      = PatternFill('solid', fgColor='EFEFEF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = _border()
    ws.row_dimensions[6].height = 20

    # Baris 7–206: Area isi data
    for row in range(7, 207):
        bg = C_WHITE if row % 2 == 0 else C_LGRAY
        for col in range(1, ncol + 1):
            c           = ws.cell(row=row, column=col, value='')
            c.border    = _border('DDDDDD')
            c.alignment = Alignment(vertical='center')
            c.fill      = PatternFill('solid', fgColor=bg)
        ws.row_dimensions[row].height = 18

    ws.freeze_panes             = 'A7'
    ws.sheet_view.showGridLines = False

    # ── Sheet 2: Referensi ────────────────────────────────────────────────────
    ws2       = wb.create_sheet('Referensi')
    ws2.merge_cells('A1:C1')
    c         = ws2['A1']
    c.value   = 'REFERENSI NILAI VALID PER KOLOM'
    c.font    = Font(bold=True, size=13, color=C_WHITE)
    c.fill    = PatternFill('solid', fgColor=C_NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    row = 3
    for field, values in REFERENSI:
        ws2.merge_cells(f'A{row}:C{row}')
        c           = ws2.cell(row=row, column=1, value=field)
        c.font      = Font(bold=True, color=C_WHITE, size=11)
        c.fill      = PatternFill('solid', fgColor=C_GREEN)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border    = _border()
        ws2.row_dimensions[row].height = 22
        row += 1
        for i, val in enumerate(values):
            c           = ws2.cell(row=row, column=2, value=val)
            c.font      = Font(size=10)
            c.fill      = PatternFill('solid', fgColor=C_WHITE if i % 2 == 0 else C_GRAY)
            c.alignment = Alignment(vertical='center', indent=1)
            c.border    = _border()
            ws2.row_dimensions[row].height = 18
            row += 1
        row += 1

    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 3
    ws2.sheet_view.showGridLines = False

    # ── Sheet 3: Petunjuk ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Petunjuk')
    PETUNJUK = [
        ('PETUNJUK IMPORT DATA ASET — HRIS SmartDesk',            True,  C_NAVY,   14),
        ('',                                                        False, None,     10),
        ('A. PERSIAPAN',                                            True,  C_GREEN,  11),
        ('1. Pastikan data Lokasi dan Vendor sudah ada di sistem (menu Asset > Lokasi / Vendor).', False, None, 10),
        ('   Jika Lokasi/Vendor tidak ditemukan, kolom tersebut akan dilewati (aset tetap masuk).',False, None, 10),
        ('2. Kategori Induk dan Kategori akan dibuat OTOMATIS jika belum ada.',                   False, None, 10),
        ('3. Penanggung Jawab: isi NIK karyawan. Jika tidak ditemukan, akan dilewati.',            False, None, 10),
        ('4. Hapus baris contoh (baris abu-abu no.6) sebelum upload.',                             False, None, 10),
        ('',                                                        False, None,     10),
        ('B. FORMAT PENTING',                                       True,  C_GREEN,  11),
        ('  Tanggal Beli / Garansi  : DD/MM/YYYY  contoh: 15/01/2023',                            False, None, 10),
        ('  Harga Beli / Nilai Residu: angka saja tanpa titik/koma  contoh: 150000000',            False, None, 10),
        ('  Masa Manfaat            : angka tahun. Isi 0 untuk tanah (tidak disusutkan)',           False, None, 10),
        ('  Kode Aset               : kosongkan untuk auto-generate sesuai kategori',              False, None, 10),
        ('',                                                        False, None,     10),
        ('C. KOLOM WAJIB (header merah, tanda *)',                  True,  C_RED,    11),
        ('  Nama Aset, Tanggal Beli, Harga Beli',                  False, None,     10),
        ('',                                                        False, None,     10),
        ('D. PROSES IMPORT',                                        True,  C_GREEN,  11),
        ('1. Isi data mulai baris ke-7 (baris 6 adalah contoh, hapus dulu)',                       False, None, 10),
        ('2. Simpan file .xlsx',                                                                    False, None, 10),
        ('3. Buka menu Asset Management > Daftar Aset > Import Aset',                             False, None, 10),
        ('4. Upload file, lihat hasil import (berhasil / error per baris)',                        False, None, 10),
        ('',                                                        False, None,     10),
        ('E. CATATAN DEPRESIASI',                                   True,  C_ORANGE, 11),
        ('  Jadwal depresiasi (garis lurus) akan digenerate OTOMATIS saat import.',               False, None, 10),
        ('  Aset dengan Masa Manfaat = 0 (misal: Tanah) tidak akan dibuat jadwal depresiasinya.', False, None, 10),
    ]

    for i, (text, bold, color, size) in enumerate(PETUNJUK, 1):
        c           = ws3.cell(row=i, column=1, value=text)
        c.font      = Font(bold=bold, size=size, color=color or '333333')
        c.alignment = Alignment(vertical='center', indent=1 if not bold else 0)
        ws3.row_dimensions[i].height = 20 if text else 8

    ws3.column_dimensions['A'].width = 90
    ws3.sheet_view.showGridLines = False

    buf  = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="template_import_aset.xlsx"'
    return resp


# ── Import processor ──────────────────────────────────────────────────────────

def import_asset_excel(file_obj, company=None):
    """
    Import data aset dari file Excel.
    Return: (success_count, error_list)
    """
    from apps.assets.models import Asset, Category as AssetCategory
    from apps.assets.depreciation import generate_depreciation
    from apps.locations.models import Location
    from apps.vendors.models import Vendor
    from apps.employees.models import Employee
    from datetime import datetime

    errors  = []
    success = 0

    try:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return 0, [{'baris': '-', 'nama': '-', 'alasan': f'File tidak bisa dibaca: {e}'}]

    # Auto-deteksi baris header
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        vals = [str(c.value or '').strip().upper() for c in row]
        if any('NAMA ASET' in v or v == 'NAMA ASET *' for v in vals):
            header_row = row[0].row
            break

    if header_row is None:
        return 0, [{'baris': '-', 'nama': '-',
                    'alasan': 'Format file tidak valid: kolom "Nama Aset" tidak ditemukan. Gunakan template resmi.'}]

    headers = []
    for cell in list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]:
        val = str(cell.value or '').strip().lower()
        val = (val.replace(' *', '').replace('/', '_').replace(' ', '_')
                  .replace('.', '').replace('(', '').replace(')', ''))
        headers.append(val)

    data_start = header_row + 3 if header_row >= 3 else header_row + 1

    def get_col(row, *fields):
        for field in fields:
            try:
                idx = headers.index(field)
                val = row[idx].value
                if val is not None:
                    if hasattr(val, 'year'):
                        return val
                    s = str(val).strip()
                    return s if s not in ('None', 'nan', '') else ''
            except (ValueError, IndexError):
                continue
        return ''

    def parse_date(s):
        if not s:
            return None
        if hasattr(s, 'date'):
            return s.date()
        if hasattr(s, 'year'):
            return s
        s = str(s).strip()
        if ' ' in s:
            s = s.split(' ')[0]
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def parse_int(s, default=0):
        try:
            return int(str(s).replace('.', '').replace(',', '').strip())
        except (ValueError, TypeError):
            return default

    def _get_or_create_category(parent_name, child_name):
        """Get or create kategori hierarki per company."""
        if not parent_name and not child_name:
            return None

        def _unique_code(name, prefix=''):
            base = (prefix + name[:4]).upper().replace(' ', '')
            code, i = base, 1
            while AssetCategory.objects.filter(code=code).exists():
                code = f'{base}{i}'
                i += 1
            return code

        parent = None
        if parent_name:
            parent, _ = AssetCategory.objects.get_or_create(
                company=company,
                name=parent_name,
                defaults={'code': _unique_code(parent_name), 'asset_type': 'Tangible', 'parent': None}
            )

        if child_name:
            child, _ = AssetCategory.objects.get_or_create(
                company=company,
                name=child_name,
                parent=parent,
                defaults={'code': _unique_code(child_name, parent.code[:2] if parent else ''), 'asset_type': 'Tangible'}
            )
            return child
        return parent

    def _gen_asset_code(prefix, company):
        """Generate kode aset unik."""
        prefix = prefix.strip().upper() if prefix else 'AST'
        last = Asset.objects.filter(
            company=company, asset_code__startswith=prefix
        ).order_by('-asset_code').first()
        start = 1
        if last:
            try:
                start = int(last.asset_code.split('-')[-1]) + 1
            except Exception:
                start = Asset.objects.filter(company=company, asset_code__startswith=prefix).count() + 1
        for n in range(start, start + 9999):
            code = f'{prefix}-{n:04d}'
            if not Asset.objects.filter(asset_code=code).exists():
                return code
        return f'{prefix}-{date.today().strftime("%Y%m%d%H%M%S")}'

    # Pre-load cache
    loc_cache  = {l.name.lower(): l for l in Location.objects.all()}
    vnd_cache  = {v.name.lower(): v for v in Vendor.objects.filter(status='Aktif')}
    emp_cache  = {e.nik: e for e in Employee.objects.filter(company=company, status='Aktif')} if company else {}

    for row_num, row in enumerate(ws.iter_rows(min_row=data_start), data_start):
        nama       = get_col(row, 'nama_aset')
        tgl_raw    = get_col(row, 'tanggal_beli')
        harga_raw  = get_col(row, 'harga_beli')

        if not nama:
            continue
        # Skip baris contoh
        if nama.lower() in ('toyota avanza 2023', 'nama aset', 'nama aset *'):
            continue

        tgl_beli = parse_date(tgl_raw)
        if not tgl_beli:
            errors.append({'baris': row_num, 'nama': nama,
                           'alasan': f'Tanggal beli "{tgl_raw}" tidak valid. Gunakan DD/MM/YYYY.'})
            continue

        harga = parse_int(harga_raw)
        if harga <= 0:
            errors.append({'baris': row_num, 'nama': nama,
                           'alasan': f'Harga beli "{harga_raw}" tidak valid. Isi angka lebih dari 0.'})
            continue

        try:
            parent_cat_name = get_col(row, 'kategori_induk')
            child_cat_name  = get_col(row, 'kategori')
            kategori        = _get_or_create_category(parent_cat_name, child_cat_name)

            kode = get_col(row, 'kode_aset')
            if not kode:
                prefix = 'AST'
                if kategori:
                    prefix = kategori.code[:6]
                kode = _gen_asset_code(prefix, company)
            elif Asset.objects.filter(asset_code=kode).exists():
                errors.append({'baris': row_num, 'nama': nama,
                               'alasan': f'Kode aset "{kode}" sudah ada di sistem.'})
                continue

            masa_manfaat  = parse_int(get_col(row, 'masa_manfaat'), 5)
            nilai_residu  = parse_int(get_col(row, 'nilai_residu'), 0)
            kondisi_raw   = get_col(row, 'kondisi') or 'Baik'
            kondisi_valid = ['Baik', 'Rusak Ringan', 'Rusak Berat', 'Dalam Perbaikan']
            kondisi       = kondisi_raw if kondisi_raw in kondisi_valid else 'Baik'

            status_raw    = (get_col(row, 'status') or 'ACTIVE').upper()
            status_valid  = ['ACTIVE', 'MAINTENANCE', 'RETIRED', 'BROKEN', 'RESERVED']
            status        = status_raw if status_raw in status_valid else 'ACTIVE'

            loc_name  = get_col(row, 'lokasi', 'location').lower()
            vnd_name  = get_col(row, 'vendor').lower()
            pic_nik   = get_col(row, 'penanggung_jawab')

            location  = loc_cache.get(loc_name)
            vendor    = vnd_cache.get(vnd_name)
            pic       = emp_cache.get(pic_nik)

            asset = Asset.objects.create(
                company        = company,
                asset_code     = kode,
                asset_name     = nama,
                category       = kategori,
                purchase_date  = tgl_beli,
                purchase_price = harga,
                useful_life    = masa_manfaat,
                residual_value = nilai_residu,
                brand          = get_col(row, 'merk__brand', 'merk_brand', 'merk'),
                serial_number  = get_col(row, 'serial_number'),
                condition      = kondisi,
                status         = status,
                location       = location,
                vendor         = vendor,
                responsible    = pic,
                warranty_date  = parse_date(get_col(row, 'garansi_sd', 'garansi')),
                invoice        = get_col(row, 'no_notainvoice', 'no_nota_invoice', 'invoice'),
                notes          = get_col(row, 'catatan'),
            )

            # Generate depresiasi otomatis
            if asset.useful_life > 0:
                try:
                    generate_depreciation(asset)
                except Exception:
                    pass  # jangan gagalkan import karena depresiasi

            success += 1

        except Exception as e:
            errors.append({'baris': row_num, 'nama': nama, 'alasan': str(e)})

    return success, errors
