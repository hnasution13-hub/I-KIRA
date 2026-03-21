"""
Holiday Management — List, CRUD, Import Excel, Download Template.
"""
import io
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from apps.attendance.models import Holiday
from apps.core.decorators import hr_required


# ── Helpers ───────────────────────────────────────────────────────────────────

def _company(request):
    return getattr(request, 'company', None)


def _parse_tanggal(val):
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    if not val:
        return None
    val = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%Y/%m/%d'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_tipe(val):
    if not val:
        return 'Nasional'
    val = str(val).strip().title()
    mapping = {
        'Nasional': 'Nasional', 'National': 'Nasional', 'Skb': 'Nasional',
        'Bersama': 'Bersama', 'Cuti Bersama': 'Bersama',
        'Internal': 'Internal', 'Perusahaan': 'Internal', 'Kebijakan': 'Internal',
    }
    return mapping.get(val, 'Nasional')


TIPE_BADGE = {
    'Nasional': 'primary',
    'Bersama':  'warning',
    'Internal': 'danger',
}


# ── List ──────────────────────────────────────────────────────────────────────

@login_required
def holiday_list(request):
    company = _company(request)
    qs = Holiday.objects.filter(company=company) if company else Holiday.objects.all()

    # Filter tahun
    tahun_list = sorted(set(qs.values_list('tanggal__year', flat=True)), reverse=True)
    tahun_now  = date.today().year
    tahun_sel  = int(request.GET.get('tahun', tahun_now))
    tipe_sel   = request.GET.get('tipe', '')

    qs = qs.filter(tanggal__year=tahun_sel)
    if tipe_sel:
        qs = qs.filter(tipe=tipe_sel)

    holidays = qs.order_by('tanggal')

    return render(request, 'attendance/holiday_list.html', {
        'holidays'   : holidays,
        'tahun_list' : tahun_list if tahun_list else [tahun_now],
        'tahun_sel'  : tahun_sel,
        'tipe_sel'   : tipe_sel,
        'tipe_choices': Holiday.TIPE_CHOICES,
        'tipe_badge' : TIPE_BADGE,
        'tahun_now'  : tahun_now,
    })


# ── Form (Tambah / Edit) ──────────────────────────────────────────────────────

@login_required
@hr_required
def holiday_form(request, pk=None):
    company  = _company(request)
    instance = get_object_or_404(Holiday, pk=pk, company=company) if pk else None

    if request.method == 'POST':
        nama    = request.POST.get('nama', '').strip()
        tanggal = _parse_tanggal(request.POST.get('tanggal', ''))
        tipe    = request.POST.get('tipe', 'Nasional')
        ket     = request.POST.get('keterangan', '').strip()

        if not nama or not tanggal:
            messages.error(request, 'Nama dan tanggal wajib diisi.')
            return render(request, 'attendance/holiday_form.html', {
                'instance': instance, 'tipe_choices': Holiday.TIPE_CHOICES,
            })

        # Cek duplikat tanggal (exclude instance saat edit)
        dup_qs = Holiday.objects.filter(company=company, tanggal=tanggal)
        if instance:
            dup_qs = dup_qs.exclude(pk=instance.pk)
        if dup_qs.exists():
            messages.error(request, f'Tanggal {tanggal.strftime("%d/%m/%Y")} sudah ada.')
            return render(request, 'attendance/holiday_form.html', {
                'instance': instance, 'tipe_choices': Holiday.TIPE_CHOICES,
            })

        if instance:
            instance.nama       = nama
            instance.tanggal    = tanggal
            instance.tipe       = tipe
            instance.keterangan = ket
            instance.save()
            messages.success(request, f'Hari libur "{nama}" berhasil diperbarui.')
        else:
            Holiday.objects.create(
                company=company, nama=nama, tanggal=tanggal,
                tipe=tipe, keterangan=ket,
            )
            messages.success(request, f'Hari libur "{nama}" berhasil ditambahkan.')
        return redirect('holiday_list')

    return render(request, 'attendance/holiday_form.html', {
        'instance'    : instance,
        'tipe_choices': Holiday.TIPE_CHOICES,
    })


# ── Delete ────────────────────────────────────────────────────────────────────

@login_required
@hr_required
def holiday_delete(request, pk):
    company  = _company(request)
    instance = get_object_or_404(Holiday, pk=pk, company=company)
    if request.method == 'POST':
        nama = instance.nama
        instance.delete()
        messages.success(request, f'Hari libur "{nama}" dihapus.')
    return redirect('holiday_list')


# ── Download Template Excel ───────────────────────────────────────────────────

@login_required
def holiday_download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hari Libur'

    RED   = 'C0392B'
    WHITE = 'FFFFFF'
    GRAY  = 'F8F9FA'
    LIGHT = 'EBF5FB'

    headers = ['Tanggal *', 'Nama Hari Libur *', 'Tipe *', 'Keterangan']
    widths  = [18, 35, 20, 35]
    contoh  = [
        ['01/01/2026', 'Tahun Baru 2026',          'Nasional', ''],
        ['27/03/2026', 'Isra Miraj Nabi Muhammad',  'Nasional', 'Berdasarkan SKB 3 Menteri'],
        ['28/03/2026', 'Cuti Bersama Isra Miraj',   'Bersama',  ''],
        ['18/08/2026', 'HUT Perusahaan',             'Internal', 'Kebijakan internal perusahaan'],
    ]

    # Header row
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color=WHITE, size=10)
        c.fill      = PatternFill('solid', fgColor=RED)
        c.alignment = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='CCCCCC')
        c.border = Border(left=thin, right=thin, bottom=thin)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    # Contoh data
    for ri, row in enumerate(contoh, 2):
        fill = PatternFill('solid', fgColor=LIGHT if ri % 2 == 0 else GRAY)
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(size=9, italic=(ri == 2))
            c.fill      = fill
            c.alignment = Alignment(horizontal='center' if ci in (1, 3) else 'left',
                                    vertical='center')
            thin = Side(style='thin', color='DDDDDD')
            c.border = Border(left=thin, right=thin, bottom=thin)
        ws.row_dimensions[ri].height = 18

    # Sheet petunjuk
    ws2 = wb.create_sheet('Petunjuk')
    ws2.column_dimensions['A'].width = 65
    petunjuk = [
        ('PETUNJUK IMPORT HARI LIBUR', True),
        ('', False),
        ('Kolom Tanggal  : Format DD/MM/YYYY atau YYYY-MM-DD. Contoh: 01/01/2026', False),
        ('Kolom Nama     : Nama hari libur, wajib diisi.', False),
        ('Kolom Tipe     : Isi salah satu dari 3 pilihan berikut:', False),
        ('                  • Nasional  — Libur nasional (SKB 3 Menteri / Keppres)', False),
        ('                  • Bersama   — Cuti bersama pemerintah', False),
        ('                  • Internal  — Kebijakan perusahaan (override)', False),
        ('Kolom Keterangan: Opsional, boleh dikosongkan.', False),
        ('', False),
        ('CATATAN:', True),
        ('• Baris pertama (header) akan diabaikan secara otomatis.', False),
        ('• Tanggal duplikat akan di-skip (tidak menimpa data yang sudah ada).', False),
        ('• Gunakan tombol Preview sebelum simpan untuk verifikasi data.', False),
        ('• Tipe selain 3 pilihan di atas akan otomatis dianggap "Nasional".', False),
    ]
    for i, (text, bold) in enumerate(petunjuk, 1):
        c = ws2.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=10 if bold else 9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="template_hari_libur.xlsx"'
    return resp


# ── Import: Upload → Preview ──────────────────────────────────────────────────

@login_required
@hr_required
def holiday_import(request):
    """Step 1: Upload file Excel, parse, simpan ke session, redirect ke preview."""
    if request.method == 'POST' and 'file' in request.FILES:
        f = request.FILES['file']
        if not f.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'Format harus .xlsx atau .xls')
            return render(request, 'attendance/holiday_import.html')

        try:
            wb   = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            messages.error(request, f'Gagal membaca file: {e}')
            return render(request, 'attendance/holiday_import.html')

        if len(rows) < 2:
            messages.error(request, 'File kosong atau hanya berisi header.')
            return render(request, 'attendance/holiday_import.html')

        # Parse data (skip baris 1 = header)
        parsed = []
        errors = []
        company = _company(request)

        for i, row in enumerate(rows[1:], 2):
            if not any(row):
                continue
            tanggal_raw = row[0] if len(row) > 0 else None
            nama_raw    = row[1] if len(row) > 1 else None
            tipe_raw    = row[2] if len(row) > 2 else 'Nasional'
            ket_raw     = row[3] if len(row) > 3 else ''

            tanggal = _parse_tanggal(tanggal_raw)
            nama    = str(nama_raw).strip() if nama_raw else ''
            tipe    = _normalize_tipe(tipe_raw)
            ket     = str(ket_raw).strip() if ket_raw else ''

            if not tanggal:
                errors.append(f'Baris {i}: Tanggal "{tanggal_raw}" tidak valid — dilewati.')
                continue
            if not nama:
                errors.append(f'Baris {i}: Nama kosong — dilewati.')
                continue

            # Cek duplikat di DB
            sudah_ada = Holiday.objects.filter(company=company, tanggal=tanggal).exists()

            parsed.append({
                'baris'    : i,
                'tanggal'  : tanggal.strftime('%Y-%m-%d'),
                'tanggal_display': tanggal.strftime('%d/%m/%Y'),
                'nama'     : nama,
                'tipe'     : tipe,
                'keterangan': ket,
                'sudah_ada': sudah_ada,
            })

        if not parsed:
            messages.error(request, 'Tidak ada data valid yang bisa diproses.')
            return render(request, 'attendance/holiday_import.html')

        # Simpan ke session
        request.session['holiday_import_data'] = parsed

        for err in errors[:5]:
            messages.warning(request, err)

        return redirect('holiday_import_preview')

    return render(request, 'attendance/holiday_import.html')


# ── Import: Preview ───────────────────────────────────────────────────────────

@login_required
@hr_required
def holiday_import_preview(request):
    """Step 2: Tampilkan preview data, user bisa uncheck baris, lalu confirm."""
    parsed = request.session.get('holiday_import_data')
    if not parsed:
        messages.error(request, 'Session habis. Upload ulang file.')
        return redirect('holiday_import')

    baru    = [p for p in parsed if not p['sudah_ada']]
    duplikat = [p for p in parsed if p['sudah_ada']]

    return render(request, 'attendance/holiday_import_preview.html', {
        'parsed'    : parsed,
        'baru'      : baru,
        'duplikat'  : duplikat,
        'tipe_badge': TIPE_BADGE,
    })


# ── Import: Confirm Simpan ────────────────────────────────────────────────────

@login_required
@hr_required
def holiday_import_confirm(request):
    """Step 3: Simpan baris yang di-checklist user."""
    if request.method != 'POST':
        return redirect('holiday_import')

    parsed = request.session.get('holiday_import_data')
    if not parsed:
        messages.error(request, 'Session habis. Upload ulang file.')
        return redirect('holiday_import')

    company   = _company(request)
    overwrite = request.POST.get('overwrite') == '1'

    # Baris yang dipilih user (checkbox)
    selected = set(request.POST.getlist('selected'))

    created = skipped = updated = 0
    for item in parsed:
        if str(item['baris']) not in selected:
            continue

        tanggal = _parse_tanggal(item['tanggal'])

        if item['sudah_ada']:
            if overwrite:
                Holiday.objects.filter(company=company, tanggal=tanggal).update(
                    nama=item['nama'], tipe=item['tipe'], keterangan=item['keterangan']
                )
                updated += 1
            else:
                skipped += 1
        else:
            Holiday.objects.get_or_create(
                company=company, tanggal=tanggal,
                defaults={
                    'nama'       : item['nama'],
                    'tipe'       : item['tipe'],
                    'keterangan' : item['keterangan'],
                }
            )
            created += 1

    # Bersihkan session
    request.session.pop('holiday_import_data', None)

    if created:
        messages.success(request, f'{created} hari libur berhasil ditambahkan.')
    if updated:
        messages.success(request, f'{updated} hari libur berhasil diperbarui.')
    if skipped:
        messages.info(request, f'{skipped} baris dilewati (sudah ada, overwrite tidak dipilih).')

    return redirect('holiday_list')
