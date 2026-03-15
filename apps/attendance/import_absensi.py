"""
Import data absensi dari Excel mesin fingerprint.
Alur: Upload → Preview kolom → Mapping → Konfirmasi → Import
"""
import io
from datetime import date, datetime, time

import openpyxl
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages

from apps.employees.models import Employee
from apps.attendance.models import Attendance


# ── Template download ────────────────────────────────────────────────────────

def download_template_absensi(request):
    """Template Excel untuk import absensi manual/fingerprint."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template Absensi'

    from openpyxl.styles import Font, PatternFill, Alignment
    NAVY  = '1a4731'
    WHITE = 'FFFFFF'
    GRAY  = 'F5F5F5'

    headers = ['NIK', 'Tanggal', 'Jam Masuk', 'Jam Keluar', 'Status', 'Keterangan']
    widths   = [15, 14, 12, 12, 14, 25]
    contoh   = ['EMP001', '2024-01-15', '08:00', '17:00', 'Hadir', 'Normal']

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color=WHITE, size=10)
        c.fill      = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for i, val in enumerate(contoh, 1):
        c = ws.cell(row=2, column=i, value=val)
        c.font = Font(italic=True, size=9, color='888888')
        c.fill = PatternFill('solid', fgColor=GRAY)
        c.alignment = Alignment(horizontal='center')

    # Sheet petunjuk
    ws2 = wb.create_sheet('Petunjuk')
    petunjuk = [
        ('PETUNJUK IMPORT ABSENSI', True),
        ('', False),
        ('Kolom NIK: wajib, harus sesuai NIK karyawan di sistem', False),
        ('Kolom Tanggal: format YYYY-MM-DD atau DD/MM/YYYY', False),
        ('Kolom Jam: format HH:MM atau HH:MM:SS', False),
        ('Kolom Status: Hadir / Tidak Hadir / Izin / Sakit / Cuti / WFH', False),
        ('', False),
        ('CATATAN MAPPING KOLOM:', True),
        ('Saat upload, Anda bisa mapping kolom dari file ke kolom sistem.', False),
        ('Nama kolom di file tidak harus sama persis.', False),
        ('Contoh: kolom "ID Karyawan" di mesin bisa di-mapping ke NIK.', False),
    ]
    for i, (text, bold) in enumerate(petunjuk, 1):
        c = ws2.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=11)
        ws2.column_dimensions['A'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="template_absensi.xlsx"'
    return resp


# ── Parse helpers ────────────────────────────────────────────────────────────

def _parse_tanggal(val):
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    if not val:
        return None
    val = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d %m %Y'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_jam(val):
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if not val:
        return None
    val = str(val).strip()
    for fmt in ('%H:%M:%S', '%H:%M', '%H.%M.%S', '%H.%M'):
        try:
            return datetime.strptime(val, fmt).time()
        except ValueError:
            continue
    return None


def _normalize_status(val):
    if not val:
        return 'Hadir'
    val = str(val).strip().title()
    mapping = {
        'Hadir': 'Hadir', 'H': 'Hadir', 'Present': 'Hadir',
        'Tidak Hadir': 'Tidak Hadir', 'Absen': 'Tidak Hadir', 'A': 'Tidak Hadir', 'Absent': 'Tidak Hadir',
        'Izin': 'Izin', 'I': 'Izin', 'Permission': 'Izin',
        'Sakit': 'Sakit', 'S': 'Sakit', 'Sick': 'Sakit',
        'Cuti': 'Cuti', 'C': 'Cuti', 'Leave': 'Cuti',
        'Wfh': 'WFH', 'Work From Home': 'WFH',
        'Libur': 'Libur',
    }
    return mapping.get(val, 'Hadir')


# ── Views ────────────────────────────────────────────────────────────────────

@login_required
def attendance_import(request):
    """Step 1: Upload file Excel."""
    if request.method == 'POST' and 'file' in request.FILES:
        f = request.FILES['file']
        if not f.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Format harus .xlsx atau .xls')
            return render(request, 'attendance/import_step1.html')

        # Baca header baris pertama
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            messages.error(request, 'File kosong.')
            return render(request, 'attendance/import_step1.html')

        headers = [str(h).strip() if h else f'Kolom {i+1}' for i, h in enumerate(rows[0])]
        preview = rows[1:6]  # 5 baris preview

        # Simpan file ke session via raw bytes
        f.seek(0)
        request.session['import_file_data'] = list(rows)
        request.session['import_headers']   = headers

        return render(request, 'attendance/import_step2.html', {
            'headers': headers,
            'preview': preview,
            'system_fields': [
                ('nik',         'NIK Karyawan *'),
                ('tanggal',     'Tanggal *'),
                ('jam_masuk',   'Jam Masuk'),
                ('jam_keluar',  'Jam Keluar'),
                ('status',      'Status Kehadiran'),
                ('keterangan',  'Keterangan'),
            ],
        })

    return render(request, 'attendance/import_step1.html')


@login_required
def attendance_import_process(request):
    """Step 3: Proses import setelah mapping dikonfirmasi."""
    if request.method != 'POST':
        return redirect('attendance_import')

    rows    = request.session.get('import_file_data', [])
    headers = request.session.get('import_headers', [])

    if not rows or not headers:
        messages.error(request, 'Session habis. Upload ulang file.')
        return redirect('attendance_import')

    # Ambil mapping dari form
    mapping = {}
    for field in ['nik', 'tanggal', 'jam_masuk', 'jam_keluar', 'status', 'keterangan']:
        val = request.POST.get(f'map_{field}', '')
        if val != '':
            try:
                mapping[field] = int(val)
            except ValueError:
                pass

    if 'nik' not in mapping or 'tanggal' not in mapping:
        messages.error(request, 'Kolom NIK dan Tanggal wajib di-mapping.')
        return redirect('attendance_import')

    # Proses baris data (skip header baris 0)
    data_rows   = rows[1:]
    success     = 0
    skipped     = 0
    errors      = []
    overwrite   = request.POST.get('overwrite') == '1'

    nik_cache = {e.nik: e for e in Employee.objects.filter(status='Aktif')}

    for i, row in enumerate(data_rows, 2):
        try:
            nik     = str(row[mapping['nik']]).strip() if row[mapping['nik']] else ''
            tgl_raw = row[mapping['tanggal']] if 'tanggal' in mapping else None
            tgl     = _parse_tanggal(tgl_raw)

            if not nik or not tgl:
                skipped += 1
                continue

            emp = nik_cache.get(nik)
            if not emp:
                errors.append(f'Baris {i}: NIK "{nik}" tidak ditemukan.')
                continue

            jam_masuk  = _parse_jam(row[mapping['jam_masuk']])  if 'jam_masuk'  in mapping and row[mapping['jam_masuk']]  else None
            jam_keluar = _parse_jam(row[mapping['jam_keluar']]) if 'jam_keluar' in mapping and row[mapping['jam_keluar']] else None
            status     = _normalize_status(row[mapping['status']] if 'status' in mapping else None)
            ket        = str(row[mapping['keterangan']]).strip() if 'keterangan' in mapping and row[mapping['keterangan']] else ''

            # Hitung keterlambatan
            telat = 0
            if jam_masuk:
                normal = time(8, 0)
                if jam_masuk > normal:
                    from datetime import datetime as dt
                    delta = dt.combine(date.today(), jam_masuk) - dt.combine(date.today(), normal)
                    telat = int(delta.total_seconds() // 60)

            if overwrite:
                att, _ = Attendance.objects.update_or_create(
                    employee=emp, tanggal=tgl,
                    defaults={
                        'check_in': jam_masuk, 'check_out': jam_keluar,
                        'status': status, 'keterlambatan': telat, 'keterangan': ket,
                    }
                )
            else:
                att, created = Attendance.objects.get_or_create(
                    employee=emp, tanggal=tgl,
                    defaults={
                        'check_in': jam_masuk, 'check_out': jam_keluar,
                        'status': status, 'keterlambatan': telat, 'keterangan': ket,
                    }
                )
                if not created:
                    skipped += 1
                    continue

            success += 1

        except Exception as e:
            errors.append(f'Baris {i}: {e}')
            if len(errors) >= 20:
                errors.append('...terlalu banyak error, proses dihentikan.')
                break

    # Hapus session
    request.session.pop('import_file_data', None)
    request.session.pop('import_headers', None)

    if success:
        messages.success(request, f'{success} data absensi berhasil diimport.')
    if skipped:
        messages.info(request, f'{skipped} baris dilewati (kosong atau duplikat).')
    for err in errors:
        messages.warning(request, err)

    return redirect('attendance_list')
