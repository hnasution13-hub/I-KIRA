"""
Utilitas export/import data karyawan ke Excel.
Update: tambah kolom Perusahaan, Point of Hire, Job Site
"""
import io
from datetime import date

import openpyxl
from difflib import get_close_matches
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

C_NAVY   = '1E3A5F'
C_RED    = 'C0392B'
C_GREEN  = '1A7A4A'
C_YELLOW = 'FFF3CD'
C_WHITE  = 'FFFFFF'
C_GRAY   = 'F5F5F5'
C_LGRAY  = 'FAFAFA'


def _border(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


EXPORT_HEADERS = [
    ('NIK',                  'nik',                18),
    ('Nama Lengkap',         'nama',               30),
    ('Perusahaan',           'perusahaan',         22),
    ('Departemen',           'department',         20),
    ('Jabatan',              'jabatan',            20),
    ('Status Karyawan',      'status_karyawan',    18),
    ('Status',               'status',             12),
    ('Tanggal Masuk',        'join_date',          16),
    ('Masa Kerja',           'masa_kerja_display', 18),
    ('Point of Hire',        'point_of_hire',      18),
    ('Job Site',             'job_site',           18),
    ('Jenis Kelamin',        'jenis_kelamin',      14),
    ('Tempat Lahir',         'tempat_lahir',       18),
    ('Tanggal Lahir',        'tanggal_lahir',      16),
    ('Agama',                'agama',              12),
    ('Pendidikan',           'pendidikan',         12),
    ('Alamat',               'alamat',             35),
    ('RT',                   'rt',                  6),
    ('RW',                   'rw',                  6),
    ('Kelurahan',            'kelurahan',          20),
    ('Kecamatan',            'kecamatan',          20),
    ('Kabupaten/Kota',       'kabupaten',          22),
    ('Provinsi',             'provinsi',           20),
    ('Kode Pos',             'kode_pos',           10),
    ('No. KTP',              'no_ktp',             20),
    ('No. NPWP',             'no_npwp',            20),
    ('No. BPJS Kesehatan',   'no_bpjs_kes',        22),
    ('No. BPJS TK',          'no_bpjs_tk',         22),
    ('Email',                'email',              28),
    ('No. HP',               'no_hp',              16),
    ('Nama Bank',            'nama_bank',          16),
    ('No. Rekening',         'no_rek',             22),
    ('A/N Rekening',         'nama_rek',           24),
]


def export_karyawan_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Karyawan'

    hf  = Font(bold=True, color=C_WHITE, size=11)
    hbg = PatternFill('solid', fgColor=C_NAVY)
    ha  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alt = PatternFill('solid', fgColor='EBF2FA')

    ws.merge_cells(f'A1:{get_column_letter(len(EXPORT_HEADERS))}1')
    c = ws['A1']
    c.value     = f'Data Karyawan — Diekspor {date.today().strftime("%d/%m/%Y")}'
    c.font      = Font(bold=True, size=13, color=C_NAVY)
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    for col, (label, _, width) in enumerate(EXPORT_HEADERS, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = hf; c.fill = hbg; c.alignment = ha; c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 22

    for ri, emp in enumerate(queryset, 3):
        bg = alt if ri % 2 == 0 else None
        for col, (_, field, _) in enumerate(EXPORT_HEADERS, 1):
            val = getattr(emp, field, '')
            if hasattr(val, 'strftime'):
                val = val.strftime('%d/%m/%Y')
            elif val is None:
                val = ''
            else:
                val = str(val) if val else ''
            c = ws.cell(row=ri, column=col, value=val)
            c.alignment = Alignment(vertical='center')
            c.border = _border()
            if bg: c.fill = bg

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(EXPORT_HEADERS))}2'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="data_karyawan_{date.today()}.xlsx"'
    return resp


IMPORT_COLUMNS = [
    ('NIK',                       'Nomor Induk Karyawan, harus unik per perusahaan tenant',   True,  15, 'EMP001'),
    ('Nama Lengkap',              'Nama lengkap sesuai KTP',                                   True,  30, 'Budi Santoso'),
    ('Vendor / Outsourcing',      'Isi HANYA jika karyawan outsourcing/sub-kon. Kosongkan jika karyawan internal tenant.', False, 28, 'PT Maju Jaya'),
    ('Departemen',                'Nama departemen (harus sudah ada di sistem tenant)',         False, 20, 'Finance'),
    ('Jabatan',                   'Nama jabatan (harus sudah ada di sistem tenant)',            False, 20, 'Staff Keuangan'),
    ('Status Karyawan',           'PKWTT / PKWT / PHL / Borongan',                            False, 18, 'PKWT'),
    ('Status',                    'Aktif / Tidak Aktif / Resign / PHK / Pensiun',             False, 12, 'Aktif'),
    ('Tanggal Masuk',             'Format: DD/MM/YYYY',                                        True,  16, '01/01/2023'),
    ('Point of Hire',             'Lokasi rekrutmen/asal karyawan (harus sudah ada di sistem)', False, 18, 'Jakarta'),
    ('Job Site',                  'Lokasi penempatan kerja saat ini (harus sudah ada di sistem)', False, 18, 'Site A'),
    ('Jenis Kelamin',       'L = Laki-laki, P = Perempuan',                             False, 14, 'L'),
    ('Tempat Lahir',        'Nama kota/kabupaten tempat lahir',                          False, 18, 'Jakarta'),
    ('Tanggal Lahir',       'Format: DD/MM/YYYY',                                        False, 16, '15/08/1995'),
    ('Agama',               'Islam/Kristen/Katolik/Hindu/Buddha/Konghucu',              False, 14, 'Islam'),
    ('Pendidikan',          'SD/SMP/SMA-SMK/D1/D2/D3/D4-S1/S2/S3',                    False, 12, 'D4/S1'),
    ('Golongan Darah',      'A/B/AB/O (boleh dengan + atau -)',                         False, 12, 'O+'),
    ('Status Nikah',        'Lajang / Menikah / Cerai',                                 False, 14, 'Menikah'),
    ('PTKP',                'TK/0 / K/0 / K/1 / K/2 / K/3 dll',                       False, 10, 'K/1'),
    ('No. KTP',             '16 digit angka tanpa spasi',                               False, 20, '3174XXXXXXXXXX0001'),
    ('No. NPWP',            'Format NPWP 15-16 digit',                                  False, 20, '12.345.678.9-012.000'),
    ('No. BPJS Kesehatan',  'Nomor BPJS Kesehatan',                                     False, 22, '0001234567890'),
    ('No. BPJS TK',         'Nomor BPJS Ketenagakerjaan',                               False, 22, '13000000000000001'),
    ('Email',               'Alamat email aktif',                                        False, 28, 'budi@email.com'),
    ('No. HP',              'Nomor HP utama',                                            False, 16, '08123456789'),
    ('No. HP Alternatif',   'Nomor HP cadangan',                                         False, 16, '08987654321'),
    ('No. Darurat',         'Nomor kontak darurat',                                      False, 16, '08111222333'),
    ('Alamat (Jalan/No)',   'Jalan, nomor rumah',                                        False, 35, 'Jl. Merdeka No. 10'),
    ('RT',                  '3 digit angka',                                             False,  6, '003'),
    ('RW',                  '3 digit angka',                                             False,  6, '005'),
    ('Kelurahan',           'Nama kelurahan/desa',                                       False, 20, 'Menteng'),
    ('Kecamatan',           'Nama kecamatan',                                            False, 20, 'Menteng'),
    ('Kabupaten/Kota',      'Nama kabupaten atau kota',                                  False, 22, 'Jakarta Pusat'),
    ('Provinsi',            'Nama provinsi',                                             False, 20, 'DKI Jakarta'),
    ('Kode Pos',            '5 digit angka',                                             False, 10, '10310'),
    ('Nama Bank',           'BCA / BRI / Mandiri / BNI dll',                           False, 16, 'BCA'),
    ('No. Rekening',        'Nomor rekening bank',                                       False, 22, '1234567890'),
    ('A/N Rekening',        'Nama pemilik rekening',                                    False, 24, 'Budi Santoso'),
]

REFERENSI = [
    ('Status Karyawan', ['PKWTT (Permanen)', 'PKWT (Kontrak)', 'PHL (Pekerja Harian Lepas)', 'Borongan']),
    ('Status', ['Aktif', 'Tidak Aktif', 'Resign', 'PHK', 'Pensiun']),
    ('Jenis Kelamin', ['L (Laki-laki)', 'P (Perempuan)']),
    ('Agama', ['Islam', 'Kristen', 'Katolik', 'Hindu', 'Buddha', 'Konghucu']),
    ('Pendidikan', ['SD', 'SMP', 'SMA/SMK', 'D1', 'D2', 'D3', 'D4/S1', 'S2', 'S3']),
    ('Golongan Darah', ['A', 'B', 'AB', 'O', 'A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-']),
    ('Status Nikah', ['Lajang', 'Menikah', 'Cerai']),
    ('PTKP', ['TK/0', 'TK/1', 'TK/2', 'TK/3', 'K/0', 'K/1', 'K/2', 'K/3', 'K/I/0', 'K/I/1', 'K/I/2', 'K/I/3']),
]


def download_template_import(company=None):
    wb   = openpyxl.Workbook()
    ncol = len(IMPORT_COLUMNS)

    # ── Sheet 1: Template Import ──────────────────────────────────────────────
    ws       = wb.active
    ws.title = 'Template Import'

    # Baris 1: Judul
    ws.merge_cells(f'A1:{get_column_letter(ncol)}1')
    c           = ws['A1']
    c.value     = 'TEMPLATE IMPORT DATA KARYAWAN — i-Kira'
    c.font      = Font(bold=True, size=14, color=C_WHITE)
    c.fill      = PatternFill('solid', fgColor=C_NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # Baris 2: Info Tenant — kunci penting agar tidak rancu
    ws.merge_cells(f'A2:{get_column_letter(ncol)}2')
    c           = ws['A2']
    tenant_nama = company.nama if company else '— (tenant dipilih saat upload oleh Developer) —'
    c.value     = f'🏢  Tenant Tujuan Import: {tenant_nama}   |   Kolom "Vendor/Outsourcing" ≠ nama tenant ini'
    c.font      = Font(bold=True, size=10, color='1A5276')
    c.fill      = PatternFill('solid', fgColor='D6EAF8')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # Baris 3: Peringatan
    ws.merge_cells(f'A3:{get_column_letter(ncol)}3')
    c           = ws['A3']
    c.value     = '⚠  Kolom merah * = WAJIB  |  Format tanggal: DD/MM/YYYY  |  Hapus baris contoh (baris 6) sebelum import  |  Lihat sheet "Referensi" & "Petunjuk"'
    c.font      = Font(bold=True, size=9, color=C_RED)
    c.fill      = PatternFill('solid', fgColor=C_YELLOW)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 20

    # Baris 4: Header kolom
    for col, (label, _, wajib, width, _) in enumerate(IMPORT_COLUMNS, 1):
        display     = f'{label} *' if wajib else label
        c           = ws.cell(row=4, column=col, value=display)
        c.font      = Font(bold=True, color=C_WHITE, size=10)
        c.fill      = PatternFill('solid', fgColor=C_RED if wajib else C_NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[4].height = 30

    # Baris 5: Keterangan kolom
    for col, (_, ket, _, _, _) in enumerate(IMPORT_COLUMNS, 1):
        c           = ws.cell(row=5, column=col, value=ket)
        c.font      = Font(italic=True, size=8, color='666666')
        c.fill      = PatternFill('solid', fgColor=C_GRAY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[5].height = 40

    # Baris 6: Contoh data (abu-abu — hapus sebelum import)
    for col, (_, _, _, _, contoh) in enumerate(IMPORT_COLUMNS, 1):
        c           = ws.cell(row=6, column=col, value=contoh)
        c.font      = Font(italic=True, size=9, color='999999')
        c.fill      = PatternFill('solid', fgColor='EFEFEF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = _border()
    ws.row_dimensions[6].height = 20

    # Baris 7–106: Area isi data
    for row in range(7, 107):
        bg = C_WHITE if row % 2 == 0 else C_LGRAY
        for col in range(1, ncol + 1):
            c           = ws.cell(row=row, column=col, value='')
            c.border    = _border('DDDDDD')
            c.alignment = Alignment(vertical='center')
            c.fill      = PatternFill('solid', fgColor=bg)
        ws.row_dimensions[row].height = 18

    ws.freeze_panes            = 'A7'
    ws.sheet_view.showGridLines = False

    # ── Sheet 2: Referensi ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Referensi')
    ws2.merge_cells('A1:C1')
    c           = ws2['A1']
    c.value     = 'REFERENSI NILAI VALID PER KOLOM'
    c.font      = Font(bold=True, size=13, color=C_WHITE)
    c.fill      = PatternFill('solid', fgColor=C_NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    row = 3
    for field, values in REFERENSI:
        ws2.merge_cells(f'A{row}:C{row}')
        c           = ws2.cell(row=row, column=1, value=field)
        c.font      = Font(bold=True, color=C_WHITE, size=11)
        c.fill      = PatternFill('solid', fgColor=C_GREEN)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border    = _border()
        ws2.row_dimensions[row].height = 22
        row += 1
        for i, val in enumerate(values):
            c           = ws2.cell(row=row, column=2, value=val)
            c.font      = Font(size=10)
            c.fill      = PatternFill('solid', fgColor=C_WHITE if i % 2 == 0 else C_GRAY)
            c.alignment = Alignment(vertical='center', indent=1)
            c.border    = _border()
            ws2.row_dimensions[row].height = 18
            row += 1
        row += 1

    ws2.column_dimensions['A'].width  = 3
    ws2.column_dimensions['B'].width  = 38
    ws2.column_dimensions['C'].width  = 3
    ws2.sheet_view.showGridLines = False

    # ── Sheet 3: Petunjuk ─────────────────────────────────────────────────────
    ws3       = wb.create_sheet('Petunjuk')
    C_ORANGE  = 'E67E22'
    PETUNJUK  = [
        ('PETUNJUK PENGISIAN & IMPORT DATA KARYAWAN', True, C_NAVY, 14),
        ('', False, None, 10),
        ('KONSEP PENTING — BACA DULU', True, C_RED, 12),
        ('Aplikasi ini menggunakan sistem Multi-Tenant. Setiap perusahaan klien (Tenant) memiliki data terpisah.', False, None, 10),
        ('Template ini digunakan untuk mengimport DATA KARYAWAN milik satu Tenant tertentu.', False, None, 10),
        ('', False, None, 10),
        ('PERBEDAAN "Tenant" vs "Vendor/Outsourcing":', True, C_ORANGE, 11),
        ('  Tenant          : Perusahaan yang berlangganan i-Kira (pemilik data).', False, None, 10),
        ('                    Contoh: PT Tambang Raya, PT Maju Sejahtera', False, None, 10),
        ('  Vendor/Outsourcing: Perusahaan penyedia karyawan outsourcing/sub-kon.', False, None, 10),
        ('                    Ini BUKAN nama tenant. Kolom ini opsional, isi hanya jika', False, None, 10),
        ('                    karyawan bukan karyawan internal tenant.', False, None, 10),
        ('                    Contoh: PT Transon, PT Manpower Indonesia', False, None, 10),
        ('', False, None, 10),
        ('A. PERSIAPAN', True, C_GREEN, 11),
        ('1. Pastikan Departemen dan Jabatan sudah dibuat di sistem tenant tujuan.', False, None, 10),
        ('2. Pastikan Point of Hire dan Job Site sudah dibuat di sistem.', False, None, 10),
        ('3. Kolom Vendor/Outsourcing: isi hanya jika karyawan berasal dari perusahaan luar.', False, None, 10),
        ('   Jika vendor belum ada di sistem, akan dibuat otomatis saat import.', False, None, 10),
        ('4. Hapus baris contoh (baris abu-abu no.6) sebelum import.', False, None, 10),
        ('', False, None, 10),
        ('B. FORMAT PENTING', True, C_GREEN, 11),
        ('  Tanggal             : DD/MM/YYYY  contoh: 01/01/1990', False, None, 10),
        ('  NIK                 : Harus unik per tenant, duplikat = update data lama', False, None, 10),
        ('  Jenis Kelamin       : L atau P saja', False, None, 10),
        ('  Vendor/Outsourcing  : Kosongkan jika karyawan internal tenant', False, None, 10),
        ('  Point of Hire       : Lokasi rekrutmen/asal karyawan', False, None, 10),
        ('  Job Site            : Lokasi penempatan kerja saat ini', False, None, 10),
        ('', False, None, 10),
        ('C. KOLOM WAJIB (header merah, tanda *)', True, C_RED, 11),
        ('  NIK, Nama Lengkap, Tanggal Masuk', False, None, 10),
        ('', False, None, 10),
        ('D. PROSES IMPORT', True, C_GREEN, 11),
        ('1. Isi data mulai baris ke-7 (baris 6 adalah contoh, hapus dulu)', False, None, 10),
        ('2. Simpan file .xlsx', False, None, 10),
        ('3. Sistem > Karyawan > Import Karyawan > Upload', False, None, 10),
        ('4. Developer: pilih Tenant tujuan dari dropdown sebelum upload', False, None, 10),
        ('5. Administrator: langsung upload, data otomatis masuk ke tenant sendiri', False, None, 10),
        ('6. Lihat hasil: berhasil / error per baris', False, None, 10),
    ]

    for i, (text, bold, color, size) in enumerate(PETUNJUK, 1):
        c           = ws3.cell(row=i, column=1, value=text)
        c.font      = Font(bold=bold, size=size, color=color or '333333')
        c.alignment = Alignment(vertical='center', indent=1 if not bold else 0)
        ws3.row_dimensions[i].height = 20 if text else 8

    ws3.column_dimensions['A'].width  = 90
    ws3.sheet_view.showGridLines = False

    buf  = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="template_import_karyawan.xlsx"'
    return resp


def import_karyawan_excel(file_obj, company=None):
    """Import data karyawan dari file Excel. Return: (success_count, error_list)"""
    from apps.employees.models import Employee, PointOfHire, JobSite, Perusahaan
    from apps.core.models import Department, Position
    from apps.wilayah.models import Provinsi, Kabupaten, Kecamatan, Kelurahan
    from datetime import datetime

    errors  = []   # list of dict: {baris, nik, nama, alasan}
    success = 0

    try:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return 0, [{'baris': '-', 'nik': '-', 'nama': '-', 'alasan': f'File tidak bisa dibaca: {e}'}]

    # Auto-deteksi baris header — cari baris yang mengandung 'NIK'
    # Scan sampai baris 10 untuk toleransi berbagai format template
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        vals = [str(c.value or '').strip().upper() for c in row]
        # Toleran: cari cell yang persis 'NIK' atau 'NIK *' atau mengandung 'NIK' di posisi pertama
        if any(v in ('NIK', 'NIK *') or v.startswith('NIK') for v in vals if v):
            # Gunakan row number Excel yang sebenarnya (bukan counter enumerate)
            header_row = row[0].row
            break

    if header_row is None:
        return 0, [{'baris': '-', 'nik': '-', 'nama': '-',
                    'alasan': (
                        'Format file tidak valid: kolom NIK tidak ditemukan di 10 baris pertama. '
                        'Pastikan menggunakan template resmi dari sistem (menu Import → Download Template). '
                        'Jika file benar, pastikan header kolom pertama adalah "NIK".'
                    )}]

    headers = []
    for cell in list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]:
        val = str(cell.value or '').strip().lower()
        val = (val.replace(' *', '').replace(' (jalan/no)', '')
                  .replace('/', '_').replace(' ', '_').replace('.', ''))
        headers.append(val)

    if 'nik' not in headers:
        return 0, [{'baris': '-', 'nik': '-', 'nama': '-', 'alasan': 'Format file tidak valid. Kolom NIK tidak ditemukan. Pastikan menggunakan template dari sistem.'}]

    # Tentukan baris mulai data:
    # Template resmi sistem:
    #   Row 1: Judul, Row 2: Info Tenant, Row 3: Peringatan
    #   Row 4: Header (NIK dst) ← header_row = 4
    #   Row 5: Keterangan kolom (skip)
    #   Row 6: Contoh data (skip)
    #   Row 7+: Data sebenarnya → data_start = header_row + 3
    # Template sederhana / file custom (header di row 1 atau 2):
    #   data_start = header_row + 1 (langsung setelah header)
    if header_row >= 3:
        data_start = header_row + 3   # skip baris keterangan + contoh
    else:
        data_start = header_row + 1   # langsung setelah header

    def get_col(row, *fields):
        for field in fields:
            try:
                idx = headers.index(field)
                val = row[idx].value
                if val is not None:
                    # Kembalikan as-is untuk datetime (biar parse_date yang handle)
                    if hasattr(val, 'year'):
                        return val
                    return str(val).strip()
            except (ValueError, IndexError):
                continue
        return ''

    def parse_date(s):
        if not s: return None
        # Kalau sudah datetime/date object (dari Excel auto-convert)
        if hasattr(s, 'date'):
            return s.date()
        if hasattr(s, 'year'):
            return s
        s = str(s).strip()
        # Buang timestamp jika ada: "2015-01-10 00:00:00" → "2015-01-10"
        if ' ' in s:
            s = s.split(' ')[0]
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
            try: return datetime.strptime(s, fmt).date()
            except ValueError: continue
        return None

    def sanitize(val, field_type='text'):
        """Buang nilai yang tidak valid: 0, None, string kosong."""
        if not val: return ''
        val = str(val).strip()
        if val in ('0', '0.0', 'None', 'nan', '-'): return ''
        if field_type == 'email' and '@' not in val: return ''
        return val

    company_filter = {'company': company} if company else {}

    dept_cache = {d.nama.lower(): d for d in Department.objects.filter(**company_filter)}
    # pos_cache: key=(nama_lower, dept_id) agar jabatan match per departemen
    # fallback: nama_lower saja jika tidak ada di dept karyawan
    _pos_qs = Position.objects.filter(**company_filter)
    pos_cache_by_dept = {(p.nama.lower(), p.department_id): p for p in _pos_qs}
    pos_cache_by_name = {}
    for p in _pos_qs:
        if p.nama.lower() not in pos_cache_by_name:
            pos_cache_by_name[p.nama.lower()] = p
    poh_cache  = {p.nama.lower(): p for p in PointOfHire.objects.filter(**company_filter)}
    site_cache = {s.nama.lower(): s for s in JobSite.objects.filter(**company_filter)}
    peru_cache = {}
    prov_cache = {p.nama.lower(): p for p in Provinsi.objects.all()}
    kab_cache  = {}
    kec_cache  = {}
    kel_cache  = {}

    # Cache bank: nama lengkap + alias → objek Bank
    from apps.wilayah.models import Bank as BankModel
    bank_nama_cache  = {b.nama.lower(): b for b in BankModel.objects.all()}
    bank_alias_cache = {b.alias.lower(): b for b in BankModel.objects.all() if b.alias}

    def fuzzy_find(name, candidates, cutoff=0.8):
        """Cari nama paling mirip dari dict keys. Return value atau None."""
        if not name:
            return None
        matches = get_close_matches(name, candidates.keys(), n=1, cutoff=cutoff)
        return candidates[matches[0]] if matches else None

    def match_bank(raw):
        """Cocokkan nama bank dari Excel ke objek Bank di DB."""
        if not raw:
            return raw  # kembalikan string kosong, simpan sebagai nama_bank string
        key = raw.strip().lower()
        # Cek exact nama lengkap
        if key in bank_nama_cache:
            return bank_nama_cache[key].nama
        # Cek exact alias
        if key in bank_alias_cache:
            return bank_alias_cache[key].nama
        # Fuzzy nama lengkap
        found = fuzzy_find(key, bank_nama_cache)
        if found:
            return found.nama
        # Fuzzy alias
        found = fuzzy_find(key, bank_alias_cache, cutoff=0.7)
        if found:
            return found.nama
        # Kembalikan as-is jika tidak ketemu
        return raw.strip()

    for row_num, row in enumerate(ws.iter_rows(min_row=data_start), data_start):
        nik  = get_col(row, 'nik')
        nama = get_col(row, 'nama_lengkap', 'nama')
        join = get_col(row, 'tanggal_masuk', 'join_date')

        if not nik or not nama:
            continue
        if nik.lower() == 'emp001' and row_num <= 3:
            continue

        join_date = parse_date(join)
        if not join_date:
            errors.append({'baris': row_num, 'nik': nik, 'nama': nama,
                           'alasan': f'Tanggal masuk "{join}" tidak valid. Gunakan format DD/MM/YYYY.'})
            continue

        try:
            # Vendor/Outsourcing — auto-create jika belum ada, BUKAN nama tenant
            peru_name = get_col(row, 'vendor_/_outsourcing', 'vendor/outsourcing', 'perusahaan').lower()
            peru_obj  = None
            if peru_name:
                if peru_name not in peru_cache:
                    peru_raw = get_col(row, 'vendor_/_outsourcing', 'vendor/outsourcing', 'perusahaan')
                    lookup = {'nama__iexact': peru_raw}
                    if company:
                        lookup['company'] = company
                    peru_obj, _ = Perusahaan.objects.get_or_create(
                        **lookup,
                        defaults={'nama': peru_raw, **(({'company': company}) if company else {})}
                    )
                    peru_cache[peru_name] = peru_obj
                peru_obj = peru_cache[peru_name]

            poh_name  = get_col(row, 'point_of_hire').lower()
            site_name = get_col(row, 'job_site').lower()
            poh_obj   = poh_cache.get(poh_name)
            site_obj  = site_cache.get(site_name)

            prov_name = get_col(row, 'provinsi').lower()
            kab_name  = get_col(row, 'kabupaten_kota', 'kabupaten').lower()
            kec_name  = get_col(row, 'kecamatan').lower()
            kel_name  = get_col(row, 'kelurahan').lower()

            # ── Matching wilayah dengan fuzzy + pembanding bertingkat ──
            prov_obj = prov_cache.get(prov_name)
            if not prov_obj and prov_name:
                prov_obj = fuzzy_find(prov_name, prov_cache)

            kab_obj = None
            if kab_name:
                if kab_name not in kab_cache:
                    # Exact dulu
                    qs = Kabupaten.objects.filter(nama__iexact=kab_name)
                    if prov_obj: qs = qs.filter(provinsi=prov_obj)
                    obj = qs.first()
                    if not obj:
                        # Fuzzy — ambil semua kabupaten di provinsi (jika ada), lalu closest match
                        scope = Kabupaten.objects.filter(provinsi=prov_obj) if prov_obj else Kabupaten.objects.all()
                        scope_dict = {k.nama.lower(): k for k in scope}
                        obj = fuzzy_find(kab_name, scope_dict)
                    kab_cache[kab_name] = obj
                kab_obj = kab_cache[kab_name]

            kec_obj = None
            if kec_name:
                if kec_name not in kec_cache:
                    qs = Kecamatan.objects.filter(nama__iexact=kec_name)
                    if kab_obj: qs = qs.filter(kabupaten=kab_obj)
                    obj = qs.first()
                    if not obj:
                        scope = Kecamatan.objects.filter(kabupaten=kab_obj) if kab_obj else Kecamatan.objects.all()
                        scope_dict = {k.nama.lower(): k for k in scope}
                        obj = fuzzy_find(kec_name, scope_dict)
                        # Gunakan kecamatan yang ditemukan untuk koreksi kabupaten jika belum ada
                        if obj and not kab_obj:
                            kab_obj = obj.kabupaten
                    kec_cache[kec_name] = obj
                kec_obj = kec_cache[kec_name]

            kel_obj = None
            if kel_name:
                if kel_name not in kel_cache:
                    qs = Kelurahan.objects.filter(nama__iexact=kel_name)
                    if kec_obj: qs = qs.filter(kecamatan=kec_obj)
                    obj = qs.first()
                    if not obj:
                        scope = Kelurahan.objects.filter(kecamatan=kec_obj) if kec_obj else Kelurahan.objects.all()
                        scope_dict = {k.nama.lower(): k for k in scope}
                        obj = fuzzy_find(kel_name, scope_dict)
                    kel_cache[kel_name] = obj
                kel_obj = kel_cache[kel_name]

            dept_name = get_col(row, 'departemen').lower()
            pos_name  = get_col(row, 'jabatan').lower()
            dept_obj  = dept_cache.get(dept_name)

            # Match jabatan ke departemen yang sesuai karyawan
            # Prioritas: (nama, dept) → kalau tidak ada → buat baru di dept karyawan
            dept_id  = dept_obj.id if dept_obj else None
            pos_obj  = pos_cache_by_dept.get((pos_name, dept_id))
            if not pos_obj and pos_name:
                # Coba fallback nama saja
                fallback = pos_cache_by_name.get(pos_name)
                if fallback and fallback.department_id == dept_id:
                    pos_obj = fallback
                elif pos_name and dept_obj and company:
                    # Buat Position baru di departemen yang benar
                    level_fallback = getattr(fallback, 'level', 'Staff') if fallback else 'Staff'
                    pos_obj, _ = Position.objects.get_or_create(
                        company=company, nama=get_col(row, 'jabatan'), department=dept_obj,
                        defaults={'level': level_fallback, 'aktif': True}
                    )
                    pos_cache_by_dept[(pos_name, dept_id)] = pos_obj
                else:
                    pos_obj = fallback  # pakai fallback meski beda dept

            defaults = {
                'nama'             : nama,
                'company'          : company,
                'perusahaan'       : peru_obj,
                'department'       : dept_obj,
                'jabatan'          : pos_obj,
                'status_karyawan'  : get_col(row, 'status_karyawan') or 'PKWT',
                'status'           : get_col(row, 'status') or 'Aktif',
                'join_date'        : join_date,
                'point_of_hire'    : poh_obj,
                'job_site'         : site_obj,
                'jenis_kelamin'    : get_col(row, 'jenis_kelamin')[:1].upper() or '',
                'tanggal_lahir'    : parse_date(get_col(row, 'tanggal_lahir')),
                'agama'            : get_col(row, 'agama'),
                'pendidikan'       : get_col(row, 'pendidikan'),
                'golongan_darah'   : get_col(row, 'golongan_darah'),
                'status_nikah'     : get_col(row, 'status_nikah'),
                'ptkp'             : get_col(row, 'ptkp'),
                'alamat'           : get_col(row, 'alamat_(jalan_no)', 'alamat'),
                'rt'               : get_col(row, 'rt'),
                'rw'               : get_col(row, 'rw'),
                'provinsi'         : prov_obj,
                'kabupaten'        : kab_obj,
                'kecamatan'        : kec_obj.nama if kec_obj else '',
                'kelurahan'        : kel_obj.nama if kel_obj else '',
                'kode_pos'         : get_col(row, 'kode_pos'),
                'no_ktp'           : get_col(row, 'no_ktp'),
                'no_npwp'          : get_col(row, 'no_npwp'),
                'no_bpjs_kes'      : get_col(row, 'no_bpjs_kesehatan', 'no_bpjs_kes'),
                'no_bpjs_tk'       : get_col(row, 'no_bpjs_tk'),
                'email'            : sanitize(get_col(row, 'email'), 'email'),
                'no_hp'            : sanitize(get_col(row, 'no_hp')),
                'hp_darurat'       : sanitize(get_col(row, 'no_hp_alternatif', 'no_darurat')),
                'nama_bank'        : match_bank(get_col(row, 'nama_bank')),
                'no_rek'           : get_col(row, 'no_rekening'),
                'nama_rek'         : get_col(row, 'a_n_rekening', 'bank_account_name'),
            }

            lookup = {'nik': nik}
            if company:
                lookup['company'] = company
            emp, created = Employee.objects.get_or_create(**lookup, defaults=defaults)

            if not created:
                changed = False
                for field, new_val in defaults.items():
                    cur_val = getattr(emp, field, None)
                    # Kosong di DB → isi
                    if not cur_val:
                        setattr(emp, field, new_val)
                        changed = True
                    # Keduanya string → update hanya jika Excel lebih panjang
                    elif isinstance(cur_val, str) and isinstance(new_val, str):
                        if len(new_val) > len(cur_val):
                            setattr(emp, field, new_val)
                            changed = True
                    # Non-teks (date, FK, dll) dan DB sudah ada → skip
                if changed:
                    emp.save()
            success += 1

        except Exception as e:
            errors.append({'baris': row_num, 'nik': nik, 'nama': nama, 'alasan': str(e)})

    return success, errors