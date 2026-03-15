"""
Utilitas export/import data upah & benefit (SalaryBenefit) ke Excel.
Pola mengikuti apps/employees/export_import.py.
"""
import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.http import HttpResponse

C_NAVY   = '1E3A5F'
C_RED    = 'C0392B'
C_GREEN  = '1A7A4A'
C_YELLOW = 'FFF3CD'
C_WHITE  = 'FFFFFF'
C_GRAY   = 'F5F5F5'
C_LGRAY  = 'FAFAFA'


def _border(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _rupiah(val):
    """Format angka ke string, 0 jika kosong."""
    try:
        return int(val or 0)
    except (TypeError, ValueError):
        return 0


# ── Kolom Export ──────────────────────────────────────────────────────────────
EXPORT_HEADERS = [
    # Identitas
    ('NIK',                          'nik',                           12),
    ('Nama Karyawan',                'nama',                          28),
    ('Departemen',                   'department',                    20),
    ('Jabatan',                      'jabatan',                       20),
    # Dasar Upah
    ('Jenis Pengupahan',             'jenis_pengupahan',              16),
    ('Status Gaji',                  'status_gaji',                   14),
    ('Hari Kerja/Minggu',            'hari_kerja_per_minggu',         16),
    ('Gaji Pokok',                   'gaji_pokok',                    16),
    # Tunjangan Tetap
    ('Tunjangan Jabatan',            'tunjangan_jabatan',             16),
    ('Tunjangan Tempat Tinggal',     'tunjangan_tempat_tinggal',      20),
    ('Tunjangan Keahlian',           'tunjangan_keahlian',            16),
    ('Tunjangan Komunikasi',         'tunjangan_komunikasi',          16),
    ('Tunjangan Kesehatan',          'tunjangan_kesehatan',           16),
    # Tunjangan Tidak Tetap
    ('Tunjangan Transport',          'tunjangan_transport',           16),
    ('Tunjangan Makan',              'tunjangan_makan',               14),
    ('Tunjangan Site',               'tunjangan_site',                14),
    ('Tunjangan Kehadiran',          'tunjangan_kehadiran',           16),
    # Potongan
    ('BPJS TK Override',             'bpjs_ketenagakerjaan_override', 18),
    ('BPJS Kes Override',            'bpjs_kesehatan_override',       16),
    ('PPh21 Ditanggung Perusahaan',  'pph21_ditanggung_perusahaan',   22),
    ('Potongan Absensi/Hari',        'potongan_absensi',              18),
    ('Potongan Lainnya',             'potongan_lainnya',              16),
    # Lain
    ('THR Override',                 'thr',                           14),
    ('Bonus Tahunan',                'bonus_tahunan',                 14),
    ('Tarif Lembur/Jam Override',    'lembur_tarif_per_jam',          20),
]


def export_salary_excel(queryset):
    """Export data SalaryBenefit ke Excel. queryset = Employee queryset."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Upah & Benefit'

    hf  = Font(bold=True, color=C_WHITE, size=11)
    hbg = PatternFill('solid', fgColor=C_NAVY)
    ha  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alt = PatternFill('solid', fgColor='EBF2FA')

    # Baris 1: Judul
    ws.merge_cells(f'A1:{get_column_letter(len(EXPORT_HEADERS))}1')
    c = ws['A1']
    c.value     = f'Data Upah & Benefit — Diekspor {date.today().strftime("%d/%m/%Y")}'
    c.font      = Font(bold=True, size=13, color=C_NAVY)
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # Baris 2: Header kolom
    for col, (label, _, width) in enumerate(EXPORT_HEADERS, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = hf; c.fill = hbg; c.alignment = ha; c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 22

    # Baris 3+: Data
    for ri, emp in enumerate(queryset, 3):
        bg = alt if ri % 2 == 0 else None
        try:
            sb = emp.salary_benefit
        except Exception:
            sb = None

        for col, (_, field, _) in enumerate(EXPORT_HEADERS, 1):
            # Field dari Employee
            if field in ('nik', 'nama'):
                val = getattr(emp, field, '')
            elif field == 'department':
                val = str(emp.department) if emp.department else ''
            elif field == 'jabatan':
                val = str(emp.jabatan) if emp.jabatan else ''
            # Field dari SalaryBenefit
            elif sb is None:
                val = ''
            elif field == 'pph21_ditanggung_perusahaan':
                val = 'Ya' if getattr(sb, field, False) else 'Tidak'
            else:
                val = getattr(sb, field, 0) or 0

            c = ws.cell(row=ri, column=col, value=val)
            c.alignment = Alignment(vertical='center')
            c.border = _border()
            if bg:
                c.fill = bg

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(EXPORT_HEADERS))}2'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="data_upah_benefit_{date.today()}.xlsx"'
    return resp


# ── Kolom Import ──────────────────────────────────────────────────────────────
# (label, keterangan, wajib, width, contoh)
IMPORT_COLUMNS = [
    ('NIK',                         'NIK karyawan — harus sudah ada di sistem',                                  True,  12, 'EMP001'),
    ('Jenis Pengupahan',            'bulanan / mingguan / harian',                                               False, 16, 'bulanan'),
    ('Status Gaji',                 'reguler / all_in — all_in: lembur tidak dihitung',                          False, 16, 'reguler'),
    ('Hari Kerja/Minggu',           '5 (Senin–Jumat) atau 6 (Senin–Sabtu)',                                      False, 16, '5'),
    ('Gaji Pokok',                  'Angka tanpa titik/koma. Contoh: 5000000',                                   False, 16, '5000000'),
    ('Tunjangan Jabatan',           'Angka, 0 jika tidak ada',                                                   False, 16, '500000'),
    ('Tunjangan Tempat Tinggal',    'Angka, 0 jika tidak ada',                                                   False, 20, '0'),
    ('Tunjangan Keahlian',          'Angka, 0 jika tidak ada',                                                   False, 16, '0'),
    ('Tunjangan Komunikasi',        'Angka, 0 jika tidak ada',                                                   False, 16, '100000'),
    ('Tunjangan Kesehatan',         'Angka, 0 jika tidak ada',                                                   False, 16, '0'),
    ('Tunjangan Transport',         'Angka, 0 jika tidak ada',                                                   False, 16, '200000'),
    ('Tunjangan Makan',             'Angka, 0 jika tidak ada',                                                   False, 14, '150000'),
    ('Tunjangan Site',              'Angka, 0 jika tidak ada',                                                   False, 14, '0'),
    ('Tunjangan Kehadiran',         'Angka, 0 jika tidak ada',                                                   False, 16, '0'),
    ('BPJS TK Override',            'Angka. Isi 0 untuk hitung otomatis (3% gaji pokok)',                        False, 18, '0'),
    ('BPJS Kes Override',           'Angka. Isi 0 untuk hitung otomatis (1% gaji pokok, maks 12jt)',             False, 16, '0'),
    ('PPh21 Ditanggung Perusahaan', 'Ya / Tidak',                                                                False, 22, 'Tidak'),
    ('Potongan Absensi/Hari',       'Potongan per hari absen (Rp). Isi 0 untuk hitung otomatis',                 False, 18, '0'),
    ('Potongan Lainnya',            'Potongan tetap lain per bulan (Rp)',                                        False, 16, '0'),
    ('THR Override',                'Angka. Isi 0 untuk hitung otomatis berdasarkan masa kerja',                 False, 14, '0'),
    ('Bonus Tahunan',               'Angka, 0 jika tidak ada',                                                   False, 14, '0'),
    ('Tarif Lembur/Jam Override',   'Angka. Isi 0 untuk hitung otomatis (gaji pokok / 173)',                     False, 22, '0'),
]

REFERENSI_SALARY = [
    ('Jenis Pengupahan', ['bulanan', 'mingguan', 'harian']),
    ('Status Gaji', ['reguler — lembur dihitung normal', 'all_in — gaji sudah mencakup semua, lembur tidak dihitung']),
    ('Hari Kerja/Minggu', ['5 (Senin–Jumat)', '6 (Senin–Sabtu)']),
    ('PPh21 Ditanggung Perusahaan', ['Ya', 'Tidak']),
]


def download_template_import_salary():
    """Generate file template Excel untuk import SalaryBenefit."""
    wb = openpyxl.Workbook()
    ncol = len(IMPORT_COLUMNS)

    # Sheet 1: Template
    ws = wb.active
    ws.title = 'Template Import Gaji'

    ws.merge_cells(f'A1:{get_column_letter(ncol)}1')
    c = ws['A1']
    c.value     = 'TEMPLATE IMPORT DATA UPAH & BENEFIT — HRIS SmartDesk'
    c.font      = Font(bold=True, size=14, color=C_WHITE)
    c.fill      = PatternFill('solid', fgColor=C_NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f'A2:{get_column_letter(ncol)}2')
    c = ws['A2']
    c.value     = '⚠  Kolom merah * = WAJIB  |  Isi angka tanpa titik/koma  |  0 = otomatis/tidak ada  |  Hapus baris contoh sebelum import'
    c.font      = Font(bold=True, size=9, color=C_RED)
    c.fill      = PatternFill('solid', fgColor=C_YELLOW)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Baris 3: Header
    for col, (label, _, wajib, width, _) in enumerate(IMPORT_COLUMNS, 1):
        display = f'{label} *' if wajib else label
        c = ws.cell(row=3, column=col, value=display)
        c.font      = Font(bold=True, color=C_WHITE, size=10)
        c.fill      = PatternFill('solid', fgColor=C_RED if wajib else C_NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 35

    # Baris 4: Keterangan
    for col, (_, ket, _, _, _) in enumerate(IMPORT_COLUMNS, 1):
        c = ws.cell(row=4, column=col, value=ket)
        c.font      = Font(italic=True, size=8, color='666666')
        c.fill      = PatternFill('solid', fgColor=C_GRAY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[4].height = 45

    # Baris 5: Contoh
    for col, (_, _, _, _, contoh) in enumerate(IMPORT_COLUMNS, 1):
        c = ws.cell(row=5, column=col, value=contoh)
        c.font      = Font(italic=True, size=9, color='999999')
        c.fill      = PatternFill('solid', fgColor='EFEFEF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = _border()
    ws.row_dimensions[5].height = 20

    # Baris 6–105: Area input
    for row in range(6, 106):
        bg = C_WHITE if row % 2 == 0 else C_LGRAY
        for col in range(1, ncol + 1):
            c = ws.cell(row=row, column=col, value='')
            c.border    = _border('DDDDDD')
            c.alignment = Alignment(vertical='center')
            c.fill      = PatternFill('solid', fgColor=bg)
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = 'A6'
    ws.sheet_view.showGridLines = False

    # ── Dropdown validation ───────────────────────────────────────────────────
    # Cari index kolom untuk field yang perlu dropdown
    def col_idx(field_key):
        for i, (_, key, _, _, _) in enumerate(IMPORT_COLUMNS, 1):
            if key == field_key:
                return get_column_letter(i)
        return None

    dv_jenis = DataValidation(
        type='list', formula1='"bulanan,mingguan,harian"',
        allow_blank=True, showDropDown=False,
        error='Nilai harus: bulanan, mingguan, atau harian',
        errorTitle='Nilai tidak valid', showErrorMessage=True,
    )
    dv_status = DataValidation(
        type='list', formula1='"reguler,all_in"',
        allow_blank=True, showDropDown=False,
        error='Nilai harus: reguler atau all_in',
        errorTitle='Nilai tidak valid', showErrorMessage=True,
    )
    dv_pph21 = DataValidation(
        type='list', formula1='"Ya,Tidak"',
        allow_blank=True, showDropDown=False,
        error='Nilai harus: Ya atau Tidak',
        errorTitle='Nilai tidak valid', showErrorMessage=True,
    )
    dv_hk = DataValidation(
        type='list', formula1='"5,6"',
        allow_blank=True, showDropDown=False,
    )

    for dv, key in [
        (dv_jenis,  'jenis_pengupahan'),
        (dv_status, 'status_gaji'),
        (dv_pph21,  'pph21_ditanggung_perusahaan'),
        (dv_hk,     'hari_kerja_per_minggu'),
    ]:
        col = col_idx(key)
        if col:
            dv.sqref = f'{col}6:{col}105'
            ws.add_data_validation(dv)

    # Sheet 2: Referensi
    ws2 = wb.create_sheet('Referensi')
    ws2.merge_cells('A1:C1')
    c = ws2['A1']
    c.value     = 'REFERENSI NILAI VALID PER KOLOM'
    c.font      = Font(bold=True, size=13, color=C_WHITE)
    c.fill      = PatternFill('solid', fgColor=C_NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    row = 3
    for field, values in REFERENSI_SALARY:
        ws2.merge_cells(f'A{row}:C{row}')
        c = ws2.cell(row=row, column=1, value=field)
        c.font      = Font(bold=True, color=C_WHITE, size=11)
        c.fill      = PatternFill('solid', fgColor=C_GREEN)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border    = _border()
        ws2.row_dimensions[row].height = 22
        row += 1
        for i, val in enumerate(values):
            c = ws2.cell(row=row, column=2, value=val)
            c.font      = Font(size=10)
            c.fill      = PatternFill('solid', fgColor=C_WHITE if i % 2 == 0 else C_GRAY)
            c.alignment = Alignment(vertical='center', indent=1)
            c.border    = _border()
            ws2.row_dimensions[row].height = 18
            row += 1
        row += 1

    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 38
    ws2.column_dimensions['C'].width = 3
    ws2.sheet_view.showGridLines = False

    # Sheet 3: Petunjuk
    ws3 = wb.create_sheet('Petunjuk')
    PETUNJUK = [
        ('PETUNJUK IMPORT DATA UPAH & BENEFIT', True, C_NAVY, 14),
        ('', False, None, 10),
        ('A. PERSIAPAN', True, C_GREEN, 11),
        ('1. Pastikan data karyawan (NIK) sudah ada di sistem terlebih dahulu.', False, None, 10),
        ('2. Export data karyawan aktif dari menu Upah > Export jika perlu daftar NIK.', False, None, 10),
        ('3. Hapus baris contoh (baris abu-abu no. 5) sebelum import.', False, None, 10),
        ('', False, None, 10),
        ('B. ATURAN PENGISIAN ANGKA', True, C_GREEN, 11),
        ('  Semua kolom angka diisi tanpa titik atau koma pemisah ribuan.', False, None, 10),
        ('  Contoh benar  : 5000000', False, None, 10),
        ('  Contoh salah  : 5.000.000 atau 5,000,000', False, None, 10),
        ('  Isi 0 pada kolom override jika ingin sistem menghitung otomatis.', False, None, 10),
        ('', False, None, 10),
        ('C. KOLOM OVERRIDE (isi 0 = hitung otomatis)', True, C_GREEN, 11),
        ('  BPJS TK Override      : 0 = otomatis 3% dari gaji pokok', False, None, 10),
        ('  BPJS Kes Override     : 0 = otomatis 1% dari gaji pokok (maks Rp 12 jt)', False, None, 10),
        ('  THR Override          : 0 = otomatis berdasarkan masa kerja (PP 36/2021)', False, None, 10),
        ('  Tarif Lembur Override : 0 = otomatis gaji pokok / 173', False, None, 10),
        ('  Potongan Absensi      : 0 = otomatis upah harian', False, None, 10),
        ('', False, None, 10),
        ('D. PROSES IMPORT', True, C_GREEN, 11),
        ('1. Isi data mulai baris ke-6', False, None, 10),
        ('2. Simpan file .xlsx', False, None, 10),
        ('3. Menu Payroll > Upah & Benefit > Import Gaji > Upload file', False, None, 10),
        ('4. Sistem akan UPDATE data jika NIK sudah ada, atau buat baru.', False, None, 10),
        ('5. Lihat hasil: berhasil / error per baris', False, None, 10),
    ]

    for i, (text, bold, color, size) in enumerate(PETUNJUK, 1):
        c = ws3.cell(row=i, column=1, value=text)
        c.font      = Font(bold=bold, size=size, color=color or '333333')
        c.alignment = Alignment(vertical='center', indent=1 if not bold else 0)
        ws3.row_dimensions[i].height = 20 if text else 8

    ws3.column_dimensions['A'].width = 80
    ws3.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="template_import_upah_benefit.xlsx"'
    return resp


def import_salary_excel(file_obj, company=None):
    from apps.employees.models import Employee

    errors  = []
    success = 0

    try:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return 0, [{'baris': '-', 'nik': '-', 'alasan': f'File tidak bisa dibaca: {e}'}]

    # Deteksi baris header — scan sampai row 10, gunakan row number Excel sebenarnya
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        vals = [str(c.value or '').strip().upper() for c in row]
        if any('NIK' in v for v in vals if v):
            header_row = row[0].row
            break

    if header_row is None:
        return 0, [{'baris': '-', 'nik': '-',
                    'alasan': 'Format file tidak valid. Kolom NIK tidak ditemukan di 10 baris pertama.'}]

    headers = []
    for cell in list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]:
        val = str(cell.value or '').strip().lower()
        val = val.replace(' *', '').replace('/', '_').replace(' ', '_').replace('.', '').replace('–', '-')
        headers.append(val)

    # Deteksi data_start secara dinamis:
    # Cek apakah baris tepat setelah header sudah berisi NIK (angka/kode karyawan)
    # Jika ya → data mulai header_row + 1 (template ringkas / template salary)
    # Jika tidak → skip 2 baris (baris keterangan + contoh) → header_row + 3 (template karyawan lengkap)
    nik_col_idx = None
    try:
        nik_col_idx = headers.index('nik')
    except ValueError:
        pass

    data_start = header_row + 1  # default: langsung setelah header
    if nik_col_idx is not None:
        # Ambil nilai baris header+1 di kolom NIK
        next_rows = list(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1))
        if next_rows:
            next_nik = str(next_rows[0][nik_col_idx].value or '').strip()
            # Jika kosong atau berisi teks keterangan (bukan NIK numerik/alfanumerik karyawan)
            # → kemungkinan baris keterangan/contoh, skip lebih jauh
            import re
            is_data_row = bool(re.match(r'^\d{5,}', next_nik))  # NIK biasanya angka ≥5 digit
            if not is_data_row and header_row >= 3:
                data_start = header_row + 3  # skip baris keterangan + contoh


    def get_col(row, *fields):
        for field in fields:
            try:
                idx = headers.index(field)
                val = row[idx].value
                if val is not None:
                    return str(val).strip()
            except (ValueError, IndexError):
                continue
        return ''

    def to_int(val):
        try:
            v = str(val or '').replace('.', '').replace(',', '').strip()
            return int(float(v)) if v else 0
        except (ValueError, TypeError):
            return 0

    def to_bool(val):
        return str(val or '').strip().lower() in ('ya', 'yes', '1', 'true')

    def to_decimal(val):
        from decimal import Decimal
        try:
            return Decimal(str(val or '0').replace(',', '.').strip())
        except Exception:
            return Decimal('0')

    # Cache employee — filter by company (multi-tenant)
    emp_filter = {'status': 'Aktif'}
    if company:
        emp_filter['company'] = company
    emp_cache = {e.nik: e for e in Employee.objects.filter(**emp_filter)}

    from .models import SalaryBenefit

    for row_num, row in enumerate(ws.iter_rows(min_row=data_start), data_start):
        nik = get_col(row, 'nik')
        if not nik or nik.lower() == 'emp001':
            continue

        emp = emp_cache.get(nik)
        if not emp:
            errors.append({'baris': row_num, 'nik': nik,
                           'alasan': f'NIK "{nik}" tidak ditemukan di database karyawan aktif.'})
            continue

        try:
            jenis = get_col(row, 'jenis_pengupahan') or 'bulanan'
            if jenis not in ('bulanan', 'mingguan', 'harian'):
                jenis = 'bulanan'

            status = get_col(row, 'status_gaji') or 'reguler'
            if status not in ('reguler', 'all_in'):
                status = 'reguler'

            hk_raw = to_int(get_col(row, 'hari_kerja_minggu', 'hari_kerja_per_minggu'))
            hari_kerja = 6 if hk_raw == 6 else 5

            pph21 = to_bool(get_col(row, 'pph21_ditanggung_perusahaan'))

            defaults = {
                'jenis_pengupahan'             : jenis,
                'status_gaji'                  : status,
                'hari_kerja_per_minggu'        : hari_kerja,
                'gaji_pokok'                   : to_int(get_col(row, 'gaji_pokok')),
                'tunjangan_jabatan'            : to_int(get_col(row, 'tunjangan_jabatan')),
                'tunjangan_tempat_tinggal'     : to_int(get_col(row, 'tunjangan_tempat_tinggal')),
                'tunjangan_keahlian'           : to_int(get_col(row, 'tunjangan_keahlian')),
                'tunjangan_komunikasi'         : to_int(get_col(row, 'tunjangan_komunikasi')),
                'tunjangan_kesehatan'          : to_int(get_col(row, 'tunjangan_kesehatan')),
                'tunjangan_transport'          : to_int(get_col(row, 'tunjangan_transport')),
                'tunjangan_makan'              : to_int(get_col(row, 'tunjangan_makan')),
                'tunjangan_site'               : to_int(get_col(row, 'tunjangan_site')),
                'tunjangan_kehadiran'          : to_int(get_col(row, 'tunjangan_kehadiran')),
                'bpjs_ketenagakerjaan_override': to_int(get_col(row, 'bpjs_tk_override', 'bpjs_ketenagakerjaan_override')),
                'bpjs_kesehatan_override'      : to_int(get_col(row, 'bpjs_kes_override', 'bpjs_kesehatan_override')),
                'pph21_ditanggung_perusahaan'  : pph21,
                'potongan_absensi'             : to_int(get_col(row, 'potongan_absensi_hari', 'potongan_absensi')),
                'potongan_lainnya'             : to_int(get_col(row, 'potongan_lainnya')),
                'thr'                          : to_int(get_col(row, 'thr_override', 'thr')),
                'bonus_tahunan'                : to_int(get_col(row, 'bonus_tahunan')),
                'lembur_tarif_per_jam'         : to_int(get_col(row, 'tarif_lembur_jam_override', 'lembur_tarif_per_jam')),
            }

            SalaryBenefit.objects.update_or_create(employee=emp, defaults=defaults)
            success += 1

        except Exception as e:
            errors.append({'baris': row_num, 'nik': nik, 'alasan': str(e)})

    return success, errors
