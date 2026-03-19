from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.utils import timezone
from django.db import models
import calendar
from datetime import date

from .models import SalaryBenefit, Payroll, PayrollDetail, SitePayrollSummary, SiteAllowanceRule
from apps.employees.models import Employee, JobSite
from apps.core.utils import get_company_qs, get_employee_related_qs
from apps.attendance.models import Attendance, Holiday
from apps.core.decorators import hr_required, manager_required
from utils.email_sender import send_payslip_email
from utils.email_sender import send_payslip_email
from utils.email_sender import send_payslip_email
from utils.email_sender import send_payslip_email
from utils.email_sender import send_payslip_email
from utils.email_sender import send_payslip_email
from utils.email_sender import send_payslip_email
from utils.payroll_calculator import PayrollCalculator


@login_required
@hr_required
def salary_list(request):
    # Filter by company (multi-tenant) + ambil semua data untuk DataTables client-side
    qs = SalaryBenefit.objects.select_related(
        'employee', 'employee__department', 'employee__jabatan'
    ).filter(
        employee__status='Aktif',
        **({'employee__company': request.company} if hasattr(request, 'company') and request.company else {})
    ).order_by('employee__nama')
    return render(request, 'payroll/salary_list.html', {'salaries': qs})


@login_required
@hr_required
def salary_form(request, employee_id=None):
    employee = get_object_or_404(Employee, pk=employee_id) if employee_id else None
    instance = None
    if employee:
        try:
            instance = SalaryBenefit.objects.get(employee=employee)
        except SalaryBenefit.DoesNotExist:
            pass

    if request.method == 'POST':
        emp_id = request.POST.get('employee_id') or employee_id
        emp = get_object_or_404(Employee, pk=emp_id)

        def intval(key):
            try: return int(request.POST.get(key, 0) or 0)
            except: return 0

        def decval(key):
            from decimal import Decimal
            try: return Decimal(request.POST.get(key, '0') or '0')
            except: return Decimal('0')

        data = {
            # Dasar Upah
            'jenis_pengupahan':       request.POST.get('jenis_pengupahan', 'bulanan'),
            'hari_kerja_per_minggu':  intval('hari_kerja_per_minggu') or 5,
            'status_gaji':            request.POST.get('status_gaji', 'reguler'),
            'gaji_pokok':             intval('gaji_pokok'),
            # Tunjangan Tetap
            'tunjangan_jabatan':         intval('tunjangan_jabatan'),
            'tunjangan_tempat_tinggal':  intval('tunjangan_tempat_tinggal'),
            'tunjangan_keahlian':        intval('tunjangan_keahlian'),
            'tunjangan_komunikasi':      intval('tunjangan_komunikasi'),
            'tunjangan_kesehatan':       intval('tunjangan_kesehatan'),
            # Tunjangan Tidak Tetap
            'tunjangan_transport':    intval('tunjangan_transport'),
            'tunjangan_makan':        intval('tunjangan_makan'),
            'tunjangan_site':         intval('tunjangan_site'),
            'tunjangan_kehadiran':    intval('tunjangan_kehadiran'),
            'tunjangan_alat_tipe':    request.POST.get('tunjangan_alat_tipe', ''),
            'tunjangan_alat_rate':    intval('tunjangan_alat_rate'),
            'lembur_tarif_per_jam':   intval('lembur_tarif_per_jam'),
            # Potongan
            'bpjs_ketenagakerjaan_override': intval('bpjs_ketenagakerjaan_override'),
            'bpjs_kesehatan_override':       intval('bpjs_kesehatan_override'),
            'pph21_ditanggung_perusahaan':   bool(request.POST.get('pph21_ditanggung_perusahaan')),
            'potongan_absensi':              intval('potongan_absensi'),
            'potongan_lainnya':              intval('potongan_lainnya'),
            # Tunjangan Lain
            'thr':          intval('thr'),
            'bonus_tahunan': intval('bonus_tahunan'),
            # Payroll Custom
            'custom_aktif':                   bool(request.POST.get('custom_aktif')),
            'custom_lembur_jam1_multiplier':   decval('custom_lembur_jam1_multiplier'),
            'custom_lembur_jam2_multiplier':   decval('custom_lembur_jam2_multiplier'),
            'custom_lembur_libur_multiplier':  decval('custom_lembur_libur_multiplier'),
            'custom_bpjs_kes_pct':             decval('custom_bpjs_kes_pct'),
            'custom_bpjs_tk_pct':              decval('custom_bpjs_tk_pct'),
            'custom_denda_telat_per_jam':      intval('custom_denda_telat_per_jam'),
            'custom_potongan_absen_per_hari':  intval('custom_potongan_absen_per_hari'),
        }
        SalaryBenefit.objects.update_or_create(employee=emp, defaults=data)
        messages.success(request, f'Data upah & benefit {emp.nama} berhasil disimpan.')
        return redirect('salary_list')

    # Estimasi THR otomatis untuk ditampilkan di form
    thr_otomatis = 0
    if instance and employee:
        thr_otomatis = PayrollCalculator.hitung_thr(employee, instance)

    context = {
        'employee':     employee,
        'instance':     instance,
        'thr_otomatis': thr_otomatis,
        'employees':    get_company_qs(Employee, request, status='Aktif').order_by('nama') if not employee else None,
    }
    return render(request, 'payroll/salary_form.html', context)


@login_required
@hr_required
def payroll_list(request):
    qs = get_company_qs(Payroll, request).order_by('-periode')
    paginator = Paginator(qs, 20)
    page = request.GET.get('page', 1)
    payrolls = paginator.get_page(page)
    return render(request, 'payroll/payroll_list.html', {'payrolls': payrolls})


@login_required
@hr_required
def payroll_generate(request):
    if request.method == 'POST':
        mode = request.POST.get('mode', 'standard')

        # ── Tentukan periode & range tanggal ──────────────────────────────
        if mode == 'custom':
            periode   = request.POST.get('periode_custom', '').strip()
            start_str = request.POST.get('start_date', '')
            end_str   = request.POST.get('end_date', '')
            if not periode or not start_str or not end_str:
                messages.error(request, 'Mode custom: isi label periode, tanggal mulai, dan tanggal selesai.')
                return redirect('payroll_generate')
            try:
                start_date = date.fromisoformat(start_str)
                end_date   = date.fromisoformat(end_str)
                if end_date < start_date:
                    raise ValueError('Tanggal selesai lebih awal dari mulai')
            except ValueError as e:
                messages.error(request, f'Tanggal tidak valid: {e}')
                return redirect('payroll_generate')
        else:
            periode = request.POST.get('periode')  # YYYY-MM
            if not periode:
                messages.error(request, 'Periode harus diisi.')
                return redirect('payroll_generate')
            try:
                year, month = map(int, periode.split('-'))
            except ValueError:
                messages.error(request, 'Format periode salah (YYYY-MM).')
                return redirect('payroll_generate')
            start_date = date(year, month, 1)
            end_date   = date(year, month, calendar.monthrange(year, month)[1])

        # Cek sudah ada
        if get_company_qs(Payroll, request).filter(periode=periode).exists():
            messages.warning(request, f'Payroll periode {periode} sudah dibuat.')
            return redirect('payroll_list')

        # Ambil hari libur nasional dari database untuk bulan ini
        holiday_dates = list(
            Holiday.objects.filter(
                tanggal__range=[start_date, end_date]
            ).values_list('tanggal', flat=True)
        )

        # Resolve company lebih awal — dipakai di dalam loop (SiteAllowanceRule)
        company = getattr(request, 'company', None)

        job_site_id = request.POST.get('job_site', '')
        employees = get_company_qs(Employee, request, status='Aktif').select_related(
            'salary_benefit', 'department', 'jabatan'
        )
        if job_site_id:
            employees = employees.filter(job_site_id=job_site_id)
        details = []
        total_kotor = 0
        total_tunjangan = 0
        total_potongan = 0
        total_bersih = 0

        for emp in employees:
            try:
                sb = emp.salary_benefit
            except SalaryBenefit.DoesNotExist:
                continue

            # Hitung hari kerja per karyawan — sesuai jadwal 5 atau 6 hari
            hari_kerja = PayrollCalculator.hitung_hari_kerja(
                start_date, end_date,
                holiday_dates=holiday_dates,
                hari_kerja_per_minggu=sb.hari_kerja_per_minggu,
            )

            # Hitung summary absensi
            attendances = Attendance.objects.filter(
                employee=emp, tanggal__range=[start_date, end_date]
            )
            hari_hadir = attendances.filter(status='Hadir').count()
            hari_absen = max(0, hari_kerja - hari_hadir - attendances.filter(
                status__in=['Cuti', 'Izin', 'Sakit']).count())
            menit_telat = sum(a.keterlambatan for a in attendances)
            jam_lembur = sum(float(a.lembur_jam) for a in attendances)

            att_summary = {
                'hari_kerja': hari_kerja,
                'hari_hadir': hari_hadir,
                'hari_absen': hari_absen,
                'menit_telat': menit_telat,
                'jam_lembur': jam_lembur,
            }

            # THR: gunakan nilai manual jika diisi, otomatis jika 0
            thr_override = sb.thr if sb.thr and sb.thr > 0 else None
            slip = PayrollCalculator.generate_slip_gaji(
                emp, sb, periode, att_summary, thr_override=thr_override
            )

            # P3.4 — Apply SiteAllowanceRule: tambahkan tunjangan site otomatis
            if emp.job_site_id:
                site_rules = SiteAllowanceRule.objects.filter(
                    company=company or emp.company,
                    job_site_id=emp.job_site_id,
                    aktif=True,
                ).filter(
                    models.Q(jabatan__isnull=True) | models.Q(jabatan=emp.jabatan)
                )
                extra_site = 0
                for rule in site_rules:
                    if rule.jenis == 'flat':
                        extra_site += rule.nilai
                    else:  # persen dari gaji pokok
                        extra_site += int(slip['gaji_pokok'] * rule.nilai / 100)
                if extra_site:
                    slip['tunjangan_site'] = slip.get('tunjangan_site', 0) + extra_site
                    slip['gaji_kotor']     = slip.get('gaji_kotor', 0) + extra_site
                    slip['gaji_bersih']    = slip.get('gaji_bersih', 0) + extra_site

            details.append(slip)
            total_kotor += slip['gaji_kotor']
            total_tunjangan += (slip['tunjangan_transport'] + slip['tunjangan_makan'] +
                                slip['tunjangan_komunikasi'] + slip['tunjangan_kesehatan'] +
                                slip['tunjangan_jabatan'] + slip['tunjangan_keahlian'] +
                                slip.get('tunjangan_site', 0))
            total_potongan += slip['total_potongan']
            total_bersih += slip['gaji_bersih']

        # Simpan payroll header
        # Developer (superuser) tidak punya request.company → ambil dari employee pertama
        if not company and employees.exists():
            company = employees.first().company
        if not company:
            messages.error(request, 'Gagal generate payroll: company tidak ditemukan.')
            return redirect('payroll_list')
        payroll = Payroll.objects.create(
            company=company,
            periode=periode,
            jumlah_karyawan=len(details),
            total_gaji_kotor=total_kotor,
            total_tunjangan=total_tunjangan,
            total_potongan=total_potongan,
            total_gaji_bersih=total_bersih,
            status='DRAFT',
        )

        # Simpan detail
        for d in details:
            PayrollDetail.objects.create(
                payroll=payroll,
                employee_id=d['employee_id'],
                gaji_pokok=d['gaji_pokok'],
                tunjangan_jabatan=d['tunjangan_jabatan'],
                tunjangan_tempat_tinggal=d['tunjangan_tempat_tinggal'],  # FIX: sebelumnya hilang
                tunjangan_keahlian=d['tunjangan_keahlian'],
                tunjangan_komunikasi=d['tunjangan_komunikasi'],
                tunjangan_kesehatan=d['tunjangan_kesehatan'],
                tunjangan_transport=d['tunjangan_transport'],
                tunjangan_makan=d['tunjangan_makan'],
                tunjangan_site=d['tunjangan_site'],           # FIX: sebelumnya hilang
                tunjangan_kehadiran=d['tunjangan_kehadiran'], # FIX: sebelumnya hilang
                upah_lembur=d['upah_lembur'],
                hari_kerja=d['hari_kerja'],
                hari_hadir=d['hari_hadir'],
                hari_absen=d['hari_absen'],
                menit_telat=d['menit_telat'],
                jam_lembur=d['jam_lembur'],
                potongan_telat=d['potongan_telat'],
                potongan_absen=d['potongan_absen'],
                potongan_lainnya=d['potongan_lainnya'],       # FIX: sebelumnya hilang
                bpjs_kesehatan=d['bpjs_kesehatan'],
                bpjs_ketenagakerjaan=d['bpjs_ketenagakerjaan'],
                pph21=d['pph21'],
                gaji_kotor=d['gaji_kotor'],
                total_potongan=d['total_potongan'],
                gaji_bersih=d['gaji_bersih'],
            )

        messages.success(request, f'Payroll {periode} berhasil dibuat untuk {len(details)} karyawan.')
        return redirect('payroll_detail', pk=payroll.pk)

    return render(request, 'payroll/payroll_generate.html', {
        'job_sites': JobSite.objects.filter(company=getattr(request, 'company', None), aktif=True),
    })


@login_required
@hr_required
def payroll_detail(request, pk):
    payroll = get_object_or_404(Payroll, pk=pk)
    details = PayrollDetail.objects.filter(payroll=payroll).select_related(
        'employee', 'employee__department', 'employee__jabatan', 'employee__job_site'
    )

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve' and payroll.status == 'DRAFT':
            payroll.status = 'APPROVED'
            payroll.approved_date = timezone.now()
            payroll.save()
            # ── Auto-generate SitePayrollSummary ───────────────────────────
            _build_site_summary(payroll, details)
            messages.success(request, 'Payroll berhasil diapprove. Ringkasan per site tersedia.')
        elif action == 'mark_paid' and payroll.status == 'APPROVED':
            payroll.status = 'PAID'
            payroll.payment_date = date.today()
            payroll.save()
            # Kirim slip gaji ke semua karyawan yang punya email
            sent, skipped = 0, 0
            for detail in details:
                if send_payslip_email(detail):
                    sent += 1
                else:
                    skipped += 1
            messages.success(request, f'Payroll ditandai sudah dibayar. Slip gaji terkirim ke {sent} karyawan ({skipped} tanpa email).')
        return redirect('payroll_detail', pk=pk)

    site_summaries = payroll.site_summaries.all()
    return render(request, 'payroll/payroll_detail.html', {
        'payroll'       : payroll,
        'details'       : details,
        'site_summaries': site_summaries,
    })


def _build_site_summary(payroll, details):
    """Hitung dan simpan SitePayrollSummary per job site dari detail payroll."""
    from collections import defaultdict
    buckets = defaultdict(lambda: {
        'job_site': None, 'site_label': '', 'rows': []
    })

    for d in details:
        site = d.employee.job_site
        key  = site.pk if site else 'none'
        buckets[key]['job_site']   = site
        buckets[key]['site_label'] = site.nama if site else 'Kantor Pusat / Tanpa Site'
        buckets[key]['rows'].append(d)

    # Hapus summary lama jika regenerate
    SitePayrollSummary.objects.filter(payroll=payroll).delete()

    for key, data in buckets.items():
        rows = data['rows']
        SitePayrollSummary.objects.create(
            payroll         = payroll,
            job_site        = data['job_site'],
            site_label      = data['site_label'],
            jumlah_karyawan = len(rows),
            total_gaji_kotor  = sum(r.gaji_kotor for r in rows),
            total_tunjangan   = sum(
                r.tunjangan_jabatan + r.tunjangan_tempat_tinggal + r.tunjangan_keahlian +
                r.tunjangan_komunikasi + r.tunjangan_kesehatan + r.tunjangan_transport +
                r.tunjangan_makan + r.tunjangan_site + r.tunjangan_kehadiran
                for r in rows
            ),
            total_potongan    = sum(r.total_potongan for r in rows),
            total_gaji_bersih = sum(r.gaji_bersih for r in rows),
        )


@login_required
@hr_required
def payroll_site_summary(request):
    """Laporan ringkasan payroll lintas periode per job site."""
    company = getattr(request, 'company', None)
    qs = Payroll.objects.filter(status__in=['APPROVED', 'PAID'])
    if company:
        qs = qs.filter(company=company)
    qs = qs.order_by('-periode')

    # Filter periode
    periode = request.GET.get('periode', '')
    if periode:
        qs = qs.filter(periode=periode)

    summaries = SitePayrollSummary.objects.filter(payroll__in=qs).select_related(
        'payroll', 'job_site'
    ).order_by('-payroll__periode', 'site_label')

    # Agregasi per site lintas periode
    from django.db.models import Sum
    site_totals = SitePayrollSummary.objects.filter(
        payroll__in=Payroll.objects.filter(company=company) if company else Payroll.objects.all()
    ).values('site_label', 'job_site__nama').annotate(
        total_kotor  = Sum('total_gaji_kotor'),
        total_bersih = Sum('total_gaji_bersih'),
        total_karyawan = Sum('jumlah_karyawan'),
    ).order_by('site_label')

    periodes = Payroll.objects.filter(status__in=['APPROVED','PAID'])
    if company:
        periodes = periodes.filter(company=company)
    periodes = periodes.values_list('periode', flat=True).distinct().order_by('-periode')

    return render(request, 'payroll/payroll_site_summary.html', {
        'summaries'   : summaries,
        'site_totals' : site_totals,
        'periodes'    : periodes,
        'selected_periode': periode,
    })


@login_required
@hr_required
def payslip(request, detail_pk):
    detail = get_object_or_404(PayrollDetail, pk=detail_pk)
    # FIX P1: Gunakan template standalone tanpa sidebar untuk cetak
    return render(request, 'payroll/payslip_print.html', {'detail': detail})

@login_required
@hr_required
def export_payroll_rekap_excel(request, pk):
    """Export rekap payroll per periode ke Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    payroll = get_object_or_404(Payroll, pk=pk)
    details = payroll.details.select_related('employee__department', 'employee__jabatan').order_by('employee__nama')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Payroll {payroll.periode}'

    # Header
    header_fill = PatternFill('solid', fgColor='C40000')
    header_font = Font(bold=True, color='FFFFFF')
    headers = [
        'NIK', 'Nama', 'Departemen', 'Jabatan',
        'Gaji Pokok', 'Tunjangan', 'Upah Lembur', 'Gaji Kotor',
        'Pot. Absen', 'Pot. Telat', 'BPJS Kes', 'BPJS TK', 'PPH21', 'Total Potongan',
        'Gaji Bersih', 'Bank', 'No. Rek', 'Nama Rek',
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    for d in details:
        emp = d.employee
        tunjangan = (d.tunjangan_jabatan or 0) + (d.tunjangan_transport or 0) +                     (d.tunjangan_makan or 0) + (d.tunjangan_komunikasi or 0) +                     (d.tunjangan_kesehatan or 0)
        ws.append([
            emp.nik, emp.nama,
            str(emp.department) if emp.department else '',
            str(emp.jabatan) if emp.jabatan else '',
            d.gaji_pokok, tunjangan, d.upah_lembur or 0, d.gaji_kotor,
            d.potongan_absen or 0, d.potongan_telat or 0,
            d.bpjs_kesehatan or 0, d.bpjs_ketenagakerjaan or 0, d.pph21 or 0,
            d.total_potongan,
            d.gaji_bersih,
            emp.nama_bank or '', emp.no_rek or '', emp.nama_rek or '',
        ])

    # Total row
    # FIX BUG-3: total_row — sebelumnya kolom tunjangan duplikat upah_lembur
    total_tunjangan_all = sum(
        (d.tunjangan_jabatan or 0) + (d.tunjangan_transport or 0) +
        (d.tunjangan_makan or 0) + (d.tunjangan_komunikasi or 0) +
        (d.tunjangan_kesehatan or 0)
        for d in details
    )
    total_row = ['', 'TOTAL', '', '',
        sum(d.gaji_pokok or 0 for d in details),
        total_tunjangan_all,
        sum(d.upah_lembur or 0 for d in details),
        sum(d.gaji_kotor or 0 for d in details),
        sum(d.potongan_absen or 0 for d in details),
        sum(d.potongan_telat or 0 for d in details),
        sum(d.bpjs_kesehatan or 0 for d in details),
        sum(d.bpjs_ketenagakerjaan or 0 for d in details),
        sum(d.pph21 or 0 for d in details),
        sum(d.total_potongan or 0 for d in details),
        sum(d.gaji_bersih or 0 for d in details),
        '', '', '',
    ]
    ws.append(total_row)
    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    # Auto width
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Payroll_{payroll.periode}.xlsx"'
    wb.save(response)
    return response


# ── Export / Import Upah & Benefit ───────────────────────────────────────────

@login_required
@hr_required
def export_salary(request):
    """Export semua data SalaryBenefit karyawan aktif ke Excel."""
    from .export_import import export_salary_excel
    qs = get_company_qs(Employee, request, status='Aktif').select_related(
        'salary_benefit', 'department', 'jabatan'
    ).order_by('nama')
    return export_salary_excel(qs)


@login_required
@hr_required
def download_template_salary(request):
    """Download template Excel kosong untuk import SalaryBenefit."""
    from .export_import import download_template_import_salary
    return download_template_import_salary()


@login_required
@hr_required
def import_salary(request):
    """Import data SalaryBenefit dari file Excel."""
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, 'Pilih file Excel terlebih dahulu.')
            return redirect('import_salary')
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Format file harus .xlsx')
            return redirect('import_salary')

        from .export_import import import_salary_excel
        company = getattr(request, 'company', None)
        success, errors = import_salary_excel(file, company=company)

        if success:
            messages.success(request, f'{success} data upah & benefit berhasil diimport/diupdate.')
        for err in errors[:20]:
            if isinstance(err, dict):
                messages.warning(request, f'Baris {err.get("baris","-")} | NIK: {err.get("nik","-")} → {err.get("alasan","-")}')
            else:
                messages.warning(request, str(err))
        if len(errors) > 20:
            messages.warning(request, f'... dan {len(errors) - 20} error lainnya (tidak ditampilkan).')
        if not success and not errors:
            messages.warning(request, 'Tidak ada data yang diimport. Pastikan file tidak kosong.')

        return redirect('salary_list')

    return render(request, 'payroll/salary_import.html')

    return render(request, 'payroll/salary_import.html')


# ─── P3.4 Site Allowance Rules ───────────────────────────────────────────────

@login_required
@hr_required
def site_allowance_list(request):
    rules = SiteAllowanceRule.objects.filter(company=request.company).select_related('job_site', 'jabatan')
    return render(request, 'payroll/site_allowance_list.html', {'rules': rules})


@login_required
@hr_required
def site_allowance_form(request, pk=None):
    from apps.core.models import Position
    instance = get_object_or_404(SiteAllowanceRule, pk=pk, company=request.company) if pk else None
    job_sites = JobSite.objects.filter(company=request.company)
    positions = Position.objects.filter(company=request.company)

    if request.method == 'POST':
        job_site_id = request.POST.get('job_site')
        jabatan_id  = request.POST.get('jabatan') or None
        nama        = request.POST.get('nama_komponen', '').strip()
        nilai       = int(request.POST.get('nilai', 0) or 0)
        jenis       = request.POST.get('jenis', 'flat')
        aktif       = request.POST.get('aktif') == 'on'

        if not job_site_id or not nama:
            messages.error(request, 'Job Site dan Nama Komponen wajib diisi.')
        else:
            if instance:
                instance.job_site_id   = job_site_id
                instance.jabatan_id    = jabatan_id
                instance.nama_komponen = nama
                instance.nilai         = nilai
                instance.jenis         = jenis
                instance.aktif         = aktif
                instance.save()
                messages.success(request, 'Aturan tunjangan site berhasil diupdate.')
            else:
                SiteAllowanceRule.objects.create(
                    company=request.company,
                    job_site_id=job_site_id,
                    jabatan_id=jabatan_id,
                    nama_komponen=nama,
                    nilai=nilai,
                    jenis=jenis,
                    aktif=aktif,
                )
                messages.success(request, 'Aturan tunjangan site berhasil ditambahkan.')
            return redirect('site_allowance_list')

    return render(request, 'payroll/site_allowance_form.html', {
        'instance': instance,
        'job_sites': job_sites,
        'positions': positions,
    })


@login_required
@hr_required
def site_allowance_delete(request, pk):
    rule = get_object_or_404(SiteAllowanceRule, pk=pk, company=request.company)
    if request.method == 'POST':
        rule.delete()
        messages.success(request, 'Aturan tunjangan site dihapus.')
        return redirect('site_allowance_list')
    return render(request, 'payroll/site_allowance_confirm_delete.html', {'rule': rule})
