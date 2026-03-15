from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from datetime import date
import io


def _get_filter_params(request):
    """Helper: ambil parameter filter bulan/tahun/departemen dari request."""
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    dept_id = request.GET.get('department', '')
    return month, year, dept_id


def _get_departments(request):
    """Helper: ambil queryset Department difilter by company (multi-tenant fix BUG-6)."""
    from apps.core.models import Department
    company = getattr(request, 'company', None)
    qs = Department.objects.filter(aktif=True)
    if company:
        qs = qs.filter(company=company)
    return qs


@login_required
def report_list(request):
    return render(request, 'reports/report_list.html')


@login_required
def report_attendance(request):
    # FIX BUG-013: Sediakan data absensi untuk laporan
    from apps.attendance.models import Attendance
    from apps.employees.models import Employee
    month, year, dept_id = _get_filter_params(request)

    qs = Attendance.objects.filter(
        tanggal__month=month, tanggal__year=year
    ).select_related('employee', 'employee__department').order_by(
        'employee__nama', 'tanggal'
    )
    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)

    return render(request, 'reports/report_attendance.html', {
        'attendances': qs,
        'departments': _get_departments(request),
        'month': month,
        'year': year,
        'selected_dept': dept_id,
        'total_records': qs.count(),
    })


@login_required
def report_payroll(request):
    # FIX BUG-013: Sediakan data payroll untuk laporan
    from apps.payroll.models import Payroll, PayrollDetail
    month, year, dept_id = _get_filter_params(request)
    periode = f"{year}-{month:02d}"

    payroll = Payroll.objects.filter(periode=periode).first()
    details = []
    if payroll:
        details = PayrollDetail.objects.filter(
            payroll=payroll
        ).select_related('employee', 'employee__department')
        if dept_id:
            details = details.filter(employee__department_id=dept_id)

    return render(request, 'reports/report_payroll.html', {
        'payroll': payroll,
        'details': details,
        'departments': _get_departments(request),
        'month': month,
        'year': year,
        'periode': periode,
        'selected_dept': dept_id,
    })


@login_required
def report_employee(request):
    # FIX BUG-013: Sediakan data karyawan untuk laporan
    from apps.employees.models import Employee
    dept_id = request.GET.get('department', '')
    status_filter = request.GET.get('status', 'Aktif')

    qs = Employee.objects.select_related('department', 'jabatan')
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if status_filter:
        qs = qs.filter(status=status_filter)

    return render(request, 'reports/report_employee.html', {
        'employees': qs,
        'departments': _get_departments(request),
        'selected_dept': dept_id,
        'status_filter': status_filter,
        'total': qs.count(),
    })


@login_required
def report_contract(request):
    # FIX BUG-013: Sediakan data kontrak untuk laporan
    from apps.contracts.models import Contract
    from datetime import timedelta
    dept_id = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')

    qs = Contract.objects.select_related('employee', 'employee__department').order_by(
        '-tanggal_mulai'
    )
    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)
    if status_filter:
        qs = qs.filter(status=status_filter)

    today = date.today()
    expiring_soon = Contract.objects.filter(
        status='Aktif',
        tanggal_selesai__isnull=False,
        tanggal_selesai__range=[today, today + timedelta(days=30)],
    ).count()

    return render(request, 'reports/report_contract.html', {
        'contracts': qs,
        'departments': _get_departments(request),
        'selected_dept': dept_id,
        'status_filter': status_filter,
        'expiring_soon': expiring_soon,
    })


@login_required
def report_violation(request):
    # FIX BUG-013: Sediakan data pelanggaran untuk laporan
    from apps.industrial.models import Violation
    month, year, dept_id = _get_filter_params(request)

    qs = Violation.objects.filter(
        tanggal_kejadian__month=month, tanggal_kejadian__year=year
    ).select_related('employee', 'employee__department').order_by('-tanggal_kejadian')

    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)

    return render(request, 'reports/report_violation.html', {
        'violations': qs,
        'departments': _get_departments(request),
        'month': month,
        'year': year,
        'selected_dept': dept_id,
        'total': qs.count(),
    })


@login_required
def report_recruitment(request):
    # FIX BUG-013: Sediakan data rekrutmen untuk laporan
    from apps.recruitment.models import Candidate, ManpowerRequest
    month, year, dept_id = _get_filter_params(request)

    candidates = Candidate.objects.filter(
        created_at__month=month, created_at__year=year
    ).select_related('mprf').order_by('-created_at')

    mprfs = ManpowerRequest.objects.filter(
        created_at__month=month, created_at__year=year
    ).select_related('department')

    if dept_id:
        mprfs = mprfs.filter(department_id=dept_id)

    return render(request, 'reports/report_recruitment.html', {
        'candidates': candidates,
        'mprfs': mprfs,
        'departments': _get_departments(request),
        'month': month,
        'year': year,
        'selected_dept': dept_id,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT FUNCTIONS — HR REPORTS
# ══════════════════════════════════════════════════════════════════════════════

def _make_excel_response(filename):
    """Helper: buat HttpResponse untuk file Excel."""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _make_pdf_response(filename):
    """Helper: buat HttpResponse untuk file PDF."""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_attendance_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl tidak terinstal.', status=500)

    from apps.attendance.models import Attendance
    month, year, dept_id = _get_filter_params(request)
    qs = Attendance.objects.filter(
        tanggal__month=month, tanggal__year=year
    ).select_related('employee', 'employee__department').order_by('employee__nama', 'tanggal')
    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Absensi'
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:J1')
    ws['A1'] = f'LAPORAN ABSENSI — {month:02d}/{year}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    headers = ['No', 'NIK', 'Nama', 'Departemen', 'Tanggal', 'Check In', 'Check Out',
               'Status', 'Keterlambatan (mnt)', 'Lembur (jam)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for idx, a in enumerate(qs, 1):
        row = idx + 3
        data = [
            idx,
            a.employee.nik,
            a.employee.nama,
            str(a.employee.department) if a.employee.department else '-',
            a.tanggal.strftime('%d/%m/%Y'),
            a.check_in.strftime('%H:%M') if a.check_in else '-',
            a.check_out.strftime('%H:%M') if a.check_out else '-',
            a.status,
            a.keterlambatan or 0,
            float(a.lembur_jam) if a.lembur_jam else 0,
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val).border = thin

    col_widths = [5, 12, 30, 20, 12, 10, 10, 12, 18, 14]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = _make_excel_response(f'laporan_absensi_{year}_{month:02d}.xlsx')
    response.write(output.read())
    return response


@login_required
def export_employee_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl tidak terinstal.', status=500)

    from apps.employees.models import Employee
    dept_id = request.GET.get('department', '')
    status_filter = request.GET.get('status', 'Aktif')

    qs = Employee.objects.select_related('department', 'jabatan')
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if status_filter:
        qs = qs.filter(status=status_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Karyawan'
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = f'DATA KARYAWAN — {date.today().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    headers = ['No', 'NIK', 'Nama', 'Departemen', 'Jabatan', 'Status',
               'Tgl Bergabung', 'Email', 'No HP']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for idx, emp in enumerate(qs, 1):
        row = idx + 3
        data = [
            idx, emp.nik, emp.nama,
            str(emp.department) if emp.department else '-',
            str(emp.jabatan) if emp.jabatan else '-',
            emp.status,
            emp.join_date.strftime('%d/%m/%Y') if emp.join_date else '-',
            emp.email or '-',
            emp.no_hp or '-',
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val).border = thin

    for i, w in enumerate([5, 12, 30, 20, 25, 10, 14, 25, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = _make_excel_response(f'laporan_karyawan_{date.today().strftime("%Y%m%d")}.xlsx')
    response.write(output.read())
    return response


@login_required
def export_payroll_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl tidak terinstal.', status=500)

    from apps.payroll.models import Payroll, PayrollDetail
    month, year, dept_id = _get_filter_params(request)
    periode = f'{year}-{month:02d}'
    payroll = Payroll.objects.filter(periode=periode).first()
    details = []
    if payroll:
        details = PayrollDetail.objects.filter(payroll=payroll).select_related('employee', 'employee__department')
        if dept_id:
            details = details.filter(employee__department_id=dept_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Payroll'
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:K1')
    ws['A1'] = f'LAPORAN PAYROLL PERIODE {periode}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    headers = ['No', 'NIK', 'Nama', 'Departemen', 'Gaji Pokok', 'Tunjangan',
               'Lembur', 'Gaji Kotor', 'Potongan', 'PPh21', 'Take Home Pay']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for idx, d in enumerate(details, 1):
        row = idx + 3
        data = [
            idx, d.employee.nik, d.employee.nama,
            str(d.employee.department) if d.employee.department else '-',
            float(d.gaji_pokok or 0),
            float((d.tunjangan_jabatan or 0) + (d.tunjangan_transport or 0) + (d.tunjangan_makan or 0)),
            float(d.upah_lembur or 0),
            float(d.gaji_kotor or 0),
            float(d.total_potongan or 0),
            float(d.pph21 or 0),
            float(d.gaji_bersih or 0),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            if col >= 5:
                cell.number_format = '#,##0'

    for i, w in enumerate([5, 12, 30, 20, 15, 15, 12, 15, 12, 12, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = _make_excel_response(f'laporan_payroll_{periode}.xlsx')
    response.write(output.read())
    return response


@login_required
def export_violation_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl tidak terinstal.', status=500)

    from apps.industrial.models import Violation
    month, year, dept_id = _get_filter_params(request)
    qs = Violation.objects.filter(
        tanggal_kejadian__month=month, tanggal_kejadian__year=year
    ).select_related('employee', 'employee__department').order_by('-tanggal_kejadian')
    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Pelanggaran'
    header_fill = PatternFill(start_color='C40000', end_color='C40000', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:H1')
    ws['A1'] = f'LAPORAN PELANGGARAN — {month:02d}/{year}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    headers = ['No', 'Nama', 'Departemen', 'Tipe SP', 'Tingkat', 'Poin', 'Tanggal', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for idx, v in enumerate(qs, 1):
        row = idx + 3
        data = [
            idx, v.employee.nama,
            str(v.employee.department) if v.employee.department else '-',
            v.tipe_pelanggaran, v.tingkat, v.poin,
            v.tanggal_kejadian.strftime('%d/%m/%Y'), v.status,
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val).border = thin

    for i, w in enumerate([5, 30, 20, 25, 10, 8, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = _make_excel_response(f'laporan_pelanggaran_{year}_{month:02d}.xlsx')
    response.write(output.read())
    return response


@login_required
def export_contract_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl tidak terinstal.', status=500)

    from apps.contracts.models import Contract
    dept_id = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')

    qs = Contract.objects.select_related('employee', 'employee__department').order_by('-tanggal_mulai')
    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)
    if status_filter:
        qs = qs.filter(status=status_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Kontrak'
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:H1')
    ws['A1'] = f'LAPORAN KONTRAK — {date.today().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    headers = ['No', 'Nama', 'Departemen', 'Tipe Kontrak', 'Tgl Mulai', 'Tgl Selesai', 'Status', 'Sisa Hari']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    today = date.today()
    for idx, c in enumerate(qs, 1):
        row = idx + 3
        sisa = (c.tanggal_selesai - today).days if c.tanggal_selesai else '-'
        data = [
            idx, c.employee.nama,
            str(c.employee.department) if c.employee.department else '-',
            c.tipe_kontrak,
            c.tanggal_mulai.strftime('%d/%m/%Y') if c.tanggal_mulai else '-',
            c.tanggal_selesai.strftime('%d/%m/%Y') if c.tanggal_selesai else 'Permanen',
            c.status, sisa,
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val).border = thin

    for i, w in enumerate([5, 30, 20, 15, 14, 14, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = _make_excel_response(f'laporan_kontrak_{date.today().strftime("%Y%m%d")}.xlsx')
    response.write(output.read())
    return response


@login_required
def export_recruitment_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl tidak terinstal.', status=500)

    from apps.recruitment.models import Candidate
    month, year, dept_id = _get_filter_params(request)
    qs = Candidate.objects.filter(
        created_at__month=month, created_at__year=year
    ).select_related('mprf').order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Rekrutmen'
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:H1')
    ws['A1'] = f'LAPORAN REKRUTMEN — {month:02d}/{year}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    headers = ['No', 'Nama Kandidat', 'Jabatan Dilamar', 'MPRF', 'Sumber',
               'Tgl Melamar', 'Status', 'Ekspektasi Gaji']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for idx, c in enumerate(qs, 1):
        row = idx + 3
        data = [
            idx, c.nama, c.jabatan_dilamar,
            c.mprf.nomor_mprf if c.mprf else '-',
            c.sumber or '-',
            c.tanggal_melamar.strftime('%d/%m/%Y') if c.tanggal_melamar else '-',
            c.status,
            float(c.ekspektasi_gaji) if c.ekspektasi_gaji else 0,
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            if col == 8:
                cell.number_format = '#,##0'

    for i, w in enumerate([5, 30, 25, 15, 15, 14, 12, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = _make_excel_response(f'laporan_rekrutmen_{year}_{month:02d}.xlsx')
    response.write(output.read())
    return response
