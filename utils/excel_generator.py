"""
Excel export generator untuk HRIS SmartDesk
Menggunakan openpyxl
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse


# Warna brand
COLOR_HEADER = "1F4E79"
COLOR_SUBHEADER = "2E75B6"
COLOR_ROW_ALT = "DEEAF1"
COLOR_TOTAL = "FFF2CC"


def style_header(cell, bold=True, bg_color=COLOR_HEADER, font_color="FFFFFF", size=11):
    cell.font = Font(bold=bold, color=font_color, size=size)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_row(cell, bg_color=None, align="left"):
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if bg_color:
        cell.fill = PatternFill("solid", fgColor=bg_color)


def thin_border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_karyawan(queryset):
    """Export data karyawan ke Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Karyawan"
    ws.row_dimensions[1].height = 30

    headers = [
        'No', 'NIK', 'Nama Lengkap', 'Departemen', 'Jabatan',
        'Status Karyawan', 'Tanggal Masuk', 'Masa Kerja', 'Status',
        'Jenis Kelamin', 'Tempat Lahir', 'Tanggal Lahir', 'Agama',
        'Pendidikan', 'No. KTP', 'No. NPWP', 'Email', 'No. HP',
        'Bank', 'No. Rekening'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(header) + 2)

    for row_idx, emp in enumerate(queryset, 2):
        bg = COLOR_ROW_ALT if row_idx % 2 == 0 else None
        data = [
            row_idx - 1, emp.nik, emp.nama,
            str(emp.department) if emp.department else '',
            str(emp.jabatan) if emp.jabatan else '',
            emp.status_karyawan,
            emp.join_date.strftime('%d/%m/%Y') if emp.join_date else '',
            emp.masa_kerja_display,
            emp.status,
            'Laki-laki' if emp.jenis_kelamin == 'L' else 'Perempuan' if emp.jenis_kelamin == 'P' else '',
            emp.tempat_lahir or '',
            emp.tanggal_lahir.strftime('%d/%m/%Y') if emp.tanggal_lahir else '',
            emp.agama or '',
            emp.pendidikan or '',
            emp.no_ktp or '',
            emp.no_npwp or '',
            emp.email or '',
            emp.no_hp or '',
            emp.bank_name or '',
            emp.bank_account or '',
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            style_row(cell, bg_color=bg, align='center' if col == 1 else 'left')
            cell.border = thin_border()

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    return _to_response(wb, 'data_karyawan.xlsx')


def export_absensi(queryset, periode_label=''):
    """Export data absensi ke Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Absensi {periode_label}"

    headers = ['No', 'NIK', 'Nama', 'Tanggal', 'Check-In', 'Check-Out',
               'Status', 'Terlambat (menit)', 'Lembur (jam)', 'Keterangan']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = max(10, len(header) + 2)

    for row_idx, att in enumerate(queryset, 2):
        bg = COLOR_ROW_ALT if row_idx % 2 == 0 else None
        data = [
            row_idx - 1, att.employee.nik, att.employee.nama,
            att.tanggal.strftime('%d/%m/%Y'),
            att.check_in.strftime('%H:%M') if att.check_in else '-',
            att.check_out.strftime('%H:%M') if att.check_out else '-',
            att.status, att.keterlambatan, float(att.lembur_jam),
            att.keterangan or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            style_row(cell, bg_color=bg)
            cell.border = thin_border()

    ws.freeze_panes = 'A2'
    return _to_response(wb, f'absensi_{periode_label}.xlsx')


def export_payroll(payroll, details):
    """Export slip gaji massal ke Excel."""
    # FIX BUG-007: Konversi ke list terlebih dahulu agar iterator tidak habis
    # sebelum len() dipanggil di bagian baris total
    details = list(details)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll {payroll.periode}"

    headers = [
        'No', 'NIK', 'Nama', 'Dept', 'Gaji Pokok', 'Tunjangan Transport',
        'Tunjangan Makan', 'Tunjangan Jabatan', 'Tunjangan Lainnya',
        'Upah Lembur', 'Gaji Kotor', 'BPJS Kes', 'BPJS TK', 'PPH21',
        'Potongan Telat', 'Potongan Absen', 'Total Potongan', 'Gaji Bersih',
        'Bank', 'No. Rekening'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = max(10, len(header) + 2)

    total_row = {h: 0 for h in range(5, 18)}

    for row_idx, d in enumerate(details, 2):
        bg = COLOR_ROW_ALT if row_idx % 2 == 0 else None
        tunjangan_lain = (
            (d.tunjangan_komunikasi or 0) + (d.tunjangan_kesehatan or 0) + (d.tunjangan_keahlian or 0)
        )
        data = [
            row_idx - 1, d.employee.nik, d.employee.nama,
            str(d.employee.department) if d.employee.department else '',
            d.gaji_pokok, d.tunjangan_transport, d.tunjangan_makan,
            d.tunjangan_jabatan, tunjangan_lain,
            d.upah_lembur, d.gaji_kotor,
            d.bpjs_kesehatan, d.bpjs_ketenagakerjaan, d.pph21,
            d.potongan_telat, d.potongan_absen, d.total_potongan,
            d.gaji_bersih,
            d.employee.nama_bank or '', d.employee.no_rek or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            style_row(cell, bg_color=bg, align='right' if col > 4 else 'left')
            if col > 4 and col < 19 and isinstance(value, (int, float)):
                cell.number_format = '#,##0'
                total_row[col] = total_row.get(col, 0) + value
            cell.border = thin_border()

    # Total row
    # FIX BUG-007: Sebelumnya len(list(details)) menghabiskan iterator dua kali → hasilnya 0
    total_row_idx = len(details) + 2
    ws.cell(row=total_row_idx, column=1, value='TOTAL')
    ws.cell(row=total_row_idx, column=3, value=f"{payroll.jumlah_karyawan} Karyawan")
    for col, total in total_row.items():
        cell = ws.cell(row=total_row_idx, column=col, value=total)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL)
        cell.number_format = '#,##0'
        cell.border = thin_border()

    ws.freeze_panes = 'A2'
    return _to_response(wb, f'payroll_{payroll.periode}.xlsx')


def _to_response(wb, filename):
    """Convert workbook to HttpResponse."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
